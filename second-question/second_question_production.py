#!/usr/bin/env python3
"""
second_question_production.py

Production helper for the source-grounded explainer video:
"The Second Question of Technology"

It does four jobs:
  1. Downloads audit/source assets and official end-card logo assets.
  2. Generates AI context clips through configurable provider profiles (Sora or
     Gemini Omni Flash), or creates local placeholder clips for editorial timing.
  3. Assembles clips deterministically against locked narration, exact on-screen
     labels, specific images, and official-logo end cards using ffmpeg.
  4. Writes metadata outputs: a release metadata workbook, flat CSVs, source
     manifest, and a DDEX ERN-style draft XML for review.

Important: the DDEX XML produced here is a draft metadata crosswalk, not a
recipient-profile-validated DDEX delivery package. DDEX recipients/distributors
have profile-specific requirements, identifiers, and validation rules.

Requirements:
  - Python 3.9+
  - ffmpeg and ffprobe available on PATH
  - pip install -r requirements_second_question.txt
  - OPENAI_API_KEY set for Sora and/or TTS generation
  - GEMINI_API_KEY set for Gemini Omni Flash context clip generation

Typical use:
  python second_question_production.py download-assets
  python second_question_production.py metadata
  python second_question_production.py generate-clips --sora
  python second_question_production.py assemble --sora --tts

Compare Sora and Gemini Omni Flash on the configured test scenes:
  python second_question_production.py compare-context-generators

Fast local mock cut without OpenAI API:
  python second_question_production.py all --mock --no-tts
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import textwrap
import time
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from xml.etree import ElementTree as ET

import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Project constants
# -----------------------------------------------------------------------------

PROJECT_TITLE = "The Second Question of Technology"
PROJECT_SLUG = "the_second_question_of_technology"
DEFAULT_ROOT = Path("second_question_build")
API_BASE = os.getenv("OPENAI_API_BASE", "https://api.openai.com/v1")

# Final edit target. AI-generated context clips may be shorter/lower resolution;
# assembly loops/trims and normalizes them deterministically to the scene timing.
DEFAULT_VIDEO_SIZE = "1920x1080"
DEFAULT_SORA_MODEL = "sora-2-pro"
DEFAULT_GEMINI_OMNI_MODEL = "gemini-omni-flash-preview"
DEFAULT_FPS = 30
END_CARD_SECONDS = 6

# Model/provider selection lives here, not in CLI flags. Edit this dictionary to
# change models, request duration, or provider-specific options. The final cut
# does not depend on generated clip duration because normalization loops/trims
# each context clip to the scripted scene length.
CONTEXT_GENERATION_PROFILES: dict[str, dict[str, Any]] = {
    "sora": {
        "provider": "sora",
        "model": DEFAULT_SORA_MODEL,
        "request_seconds": 8,
        "size": "1280x720",
        "notes": "OpenAI Videos API/Sora context visual profile.",
    },
    "gemini_omni": {
        "provider": "gemini_omni",
        "model": DEFAULT_GEMINI_OMNI_MODEL,
        "aspect_ratio": "16:9",
        "delivery": "uri",
        "notes": "Gemini Omni Flash Interactions API context visual profile.",
    },
}
DEFAULT_FULL_AI_PROFILE = "sora"
COMPARE_CONTEXT_PROFILES = ("sora", "gemini_omni")
COMPARE_CONTEXT_SCENE_IDS = (3, 5)
LOCKED_LOCAL_SCENE_IDS = {1, 8, 9}

MASTER_SORA_STYLE = """
Create a serious editorial explainer video about technological dependence and tight coupling.
The style is mechanism-focused, calm, clear, and source-grounded: original diagrams,
labeled causal chains, technical cutaway animations, archival-feeling public-domain or
Creative-Commons visual references, and abstract dependency maps.

Avoid sensational disaster montage. Avoid flames, screaming crowds, bodies, panic, or
exploitative imagery. Do not generate, redraw, imitate, or modify any logos. Leave logo-safe
blank placeholders for official logo assets to be composited later. Do not use copyrighted
movie footage or recognizable copyrighted UI. Do not invent new factual claims. Do not use
political framing. Do not include the phrase "necessity is the mother of invention." Do not say
"technology is bad" or "technology is not bad."

Use a restrained visual system: black and graphite backgrounds, white technical linework,
subtle grey panels, and one red dependency path that breaks and later becomes a resilient
network. Generate context visuals only. Exact narration, on-screen labels, logos, and
specific images will be composited in post-production. Do not render permanent text,
logos, trademarks, subtitles, captions, or UI brand names; use generic unlabeled diagram
shapes and blank callout spaces. No dialogue or narration is needed; any generated audio
will be stripped in assembly.
""".strip()

VOICEOVER_SCRIPT = """The second question in technology is the one we ask after the wreckage: why did it fail?

Challenger. Chernobyl. Fukushima. The Boeing 737 MAX crashes. Deepwater Horizon. The 2003 Northeast Blackout. Tacoma Narrows. Hyatt Regency. The Great Molasses Flood.

At Challenger, pressure escaped where the vehicle could not stay whole. At Chernobyl, cooling conditions fed back into reactor power. At Fukushima, shutdown worked, but power for cooling was trapped behind one flooded assumption. On the MAX, one angle-of-attack vane could feed repeated automated trim. On Deepwater Horizon, the last barrier could not seal. In the blackout, one disturbance became a regionwide cascade. Tacoma Narrows turned wind energy into self-excited deck motion. Hyatt Regency moved twice the load into a connection with no reserve. Boston stored a moving wall of molasses inside one failed vessel.

And sometimes no one dies, so we stop calling it disaster: the Wi-Fi goes out and the house goes dumb. A lawsuit alleges a delivery route has no bathroom slack. A refrigerator becomes an ad screen. A paid-for smart cooker becomes a subscription. A cloud shutdown turns switches into paperweights. An app becomes the boss.

These look like different failures. They are not. In each case, something became too essential, and there was no meaningful alternative path. That is tight coupling: failure stops being local because there is no meaningful substitute.

The trigger may be bad design, bad incentives, bad luck, bad maintenance, bad assumptions, or unexpected conditions. But collapse occurs when there is no bypass, reserve, repair path, manual override, interoperability, or exit.

And here is the terrifying part: experts build this. We all do. We call it efficiency, integration, convenience, optimization. But when alternatives disappear, those words become dependence.

So how do we avoid building tight coupling?

Only now do we return to the first question of technology: what need are we actually trying to satisfy?

Not the product. Not the platform. The need. Power. Communication. Transportation. Heat. Work. Safety.

Power may be grid, battery, generator, solar, or local fallback. Communication may be internet, phone, mesh, radio, or a place to meet. Transportation may be walking, bike, transit, car, rideshare, or delivery. Heat may be utility service, backup heat, insulation, or passive design. Work may be an employer, clients, cooperative platforms, local networks, and tools owned by the worker. Safety may be sensors, procedures, human override, independent checks, and fail-safe design.

A need is not reliably solved by one brittle answer. It is solved by a system of alternatives: backups, substitutes, repair, local control, open standards, interoperability, manual overrides, and exits.

The trigger of failure changes. The structural source is the same: no meaningful alternative. If the need matters, the solution must survive the failure of any single part.

If the need matters, build alternatives.
""".strip()

# -----------------------------------------------------------------------------
# Data models
# -----------------------------------------------------------------------------

@dataclass
class Source:
    key: str
    title: str
    url: str
    claim: str
    use_in_video: str
    license_note: str = "Source/audit reference; check source terms before reusing media."


@dataclass
class DownloadAsset:
    key: str
    url: str
    dest: str
    category: str
    required: bool = False
    notes: str = ""


@dataclass
class Scene:
    id: int
    slug: str
    seconds: int
    voiceover: str
    prompt: str
    labels: list[str]
    source_keys: list[str]
    reference_asset_key: Optional[str] = None

    @property
    def filename(self) -> str:
        return f"{self.id:02d}_{self.slug}.mp4"


# -----------------------------------------------------------------------------
# Source and asset manifests
# -----------------------------------------------------------------------------

SOURCES: list[Source] = [
    Source(
        "challenger",
        "NASA Rogers Commission - Challenger accident sequence",
        "https://www.nasa.gov/history/rogersrep/v1ch4.htm",
        "Right SRM aft field joint pressure seal failure and hot gas leak.",
        "Mechanism diagram: pressure seal failure -> hot gas leak -> no alternate vehicle integrity.",
    ),
    Source(
        "chernobyl",
        "OECD/NEA Chernobyl accident sequence",
        "https://oecd-nea.org/jcms/pl_28271/chernobyl-chapter-i-the-site-and-accident-sequence",
        "RBMK positive void coefficient and steam/reactivity feedback under test conditions.",
        "Feedback-loop diagram: more steam -> more reactivity -> more heat -> more steam.",
    ),
    Source(
        "fukushima",
        "TEPCO Fukushima accident sequence",
        "https://www.tepco.co.jp/en/hd/decommission/project/accident/index-e.html",
        "Earthquake shutdown followed by off-site power loss and tsunami flooding of emergency power/cooling systems.",
        "Power-chain diagram: grid, diesel, tsunami flooding, battery/cooling loss.",
    ),
    Source(
        "737max",
        "Boeing 737 MAX MCAS explanation",
        "https://www.boeing.com/content/dam/microsites/static/737-max-updates/mcas/index.html",
        "Original MCAS used one angle-of-attack input and could repeatedly command nose-down trim.",
        "Single-sensor automation diagram.",
    ),
    Source(
        "deepwater",
        "U.S. Chemical Safety Board Deepwater Horizon BOP findings",
        "https://www.csb.gov/csb-board-approves-final-report-finding-deepwater-horizon-blowout-preventer-failed-due-to-unrecognized-pipe-buckling-phenomenon-during-emergency-well-control-efforts-on-april-20-2010-leading-to-environmental-disaster-in-gulf-of-mexico/",
        "Blind shear ram did not seal buckled/off-center drill pipe; last barrier failed.",
        "Well/BOP cutaway: last barrier failed -> no fast containment path.",
    ),
    Source(
        "blackout2003",
        "U.S.-Canada Task Force 2003 Northeast Blackout report summary",
        "https://www.osti.gov/etdeweb/biblio/20461178",
        "Disturbance and line trips propagated into a broader cascade with situational-awareness failures.",
        "Grid cascade diagram.",
    ),
    Source(
        "tacoma",
        "WSDOT Tacoma Narrows Bridge failure explanation",
        "https://wsdot.wa.gov/TNBhistory/bridges-failure.htm",
        "Aeroelastic/torsional flutter; wind energy coupled into flexible deck motion.",
        "Flexible deck motion diagram; avoid simple resonance wording.",
    ),
    Source(
        "hyatt",
        "NBS/NIST Hyatt Regency collapse investigation summary",
        "https://digital.library.unt.edu/ark:/67531/metadc38324/m1/262/",
        "Hanger rod design change doubled fourth-floor box-beam connection load; no reserve capacity.",
        "Connection diagram: continuous rod vs interrupted rods.",
    ),
    Source(
        "molasses",
        "City of Boston archive - Great Molasses Flood",
        "https://www.boston.gov/news/100-years-ago-today-molasses-crashes-through-bostons-north-end",
        "Large molasses tank failed in dense neighborhood with no secondary containment.",
        "Stored hazard + failed vessel + no containment diagram.",
    ),
    Source(
        "amazon_lawsuit",
        "Cross v. Amazon complaint via FarmSTAND",
        "https://farmstand.org/wp-content/uploads/2024/04/Cross-v-Amazon-Complaint.pdf",
        "Lawsuit alleges pace and tracking prevented reasonable bathroom breaks.",
        "Route map labeled exactly: lawsuit alleges: work pace + tracking prevented breaks.",
    ),
    Source(
        "smart_fridge_ads",
        "Samsung support - Family Hub cover screen ads",
        "https://www.samsung.com/us/support/answer/ANS10007562/",
        "Support page describes ads on Family Hub cover screen and how to turn them off.",
        "Purchased appliance -> software update/policy -> ads on screen.",
    ),
    Source(
        "mellow",
        "GearBrain - Mellow sous-vide subscription coverage",
        "https://www.gearbrain.com/mellow-sous-vide-subscription-2646825555.html",
        "Owners reported smart features became subscription-gated with manual control fallback.",
        "Paid device -> server features -> subscription required -> manual mode only.",
    ),
    Source(
        "insteon",
        "WIRED - Insteon smart-home shutdown",
        "https://www.wired.com/story/insteon-shutdown/",
        "Connected switches and sensors lost server connectivity after cloud shutdown.",
        "Home control coupled to vendor survival.",
    ),
    Source(
        "uber_unsw",
        "UNSW BusinessThink - Uber algorithmic management",
        "https://businessthink.unsw.edu.au/articles/uber-algorithmic-management",
        "Research summary describes ratings, GPS monitoring, rider allocation, dynamic pricing, and chatbot support.",
        "Work coupled to opaque platform rules.",
    ),
    Source(
        "invention_image",
        "Wikimedia Commons - Angelica Kauffmann, Invention",
        "https://commons.wikimedia.org/wiki/File:Angelica_Kauffmann_(1741-1807)_-_Invention_-_03-1128_-_Royal_Academy_of_Arts.jpg",
        "Faithful reproduction of public-domain painting, used for first-question reveal.",
        "First question reveal still image.",
        "Public domain work reproduction per Commons page; verify before redistribution.",
    ),
    Source(
        "osi_brand",
        "Open Source Initiative brand and trademark guidelines",
        "https://opensource.org/about/brand-and-trademark-guidelines",
        "Use must avoid implied endorsement and follow trademark/logo usage guidelines.",
        "End-card logo asset and trademark note.",
    ),
    Source(
        "oshwa_logo",
        "Open Source Hardware Logo - OSHWA",
        "https://oshwa.org/resources/open-source-hardware-logo/",
        "Open hardware gear logo downloads and use conditions; logo may be used if hardware complies with OSHW Definition.",
        "End-card logo asset and trademark note.",
    ),
    Source(
        "beagle_brand",
        "BeagleBoard.org Brand Use",
        "https://www.beagleboard.org/brand-use",
        "Brand use rules prohibit false endorsement impressions and altered/subtly different uses.",
        "End-card logo asset and trademark note.",
    ),
    Source(
        "ddex_ern",
        "DDEX Electronic Release Notification Message Suite",
        "https://kb.ddex.net/implementing-each-standard/electronic-release-notification-message-suite-%28ern%29/",
        "ERN communicates release/resource metadata and terms to DSPs; most recent version published is ERN 4.3.2.",
        "Metadata workbook and ERN-style draft mapping.",
    ),
    Source(
        "ddex_release_delivery",
        "DDEX Release Delivery standard overview",
        "https://ddex-standards.net/standards/release-delivery/",
        "Release Delivery is defined in ERN and includes release, resources, and deal terms.",
        "Metadata workbook and ERN-style draft mapping.",
    ),
    Source(
        "openai_sora_docs",
        "OpenAI Video generation with Sora API documentation",
        "https://developers.openai.com/api/docs/guides/video-generation",
        "Sora video generation uses asynchronous video jobs and downloadable MP4 content; current docs note deprecation timing.",
        "Context video provider audit/reference.",
    ),
    Source(
        "gemini_omni_docs",
        "Google Gemini Omni Flash video generation documentation",
        "https://ai.google.dev/gemini-api/docs/omni",
        "Gemini Omni Flash uses the Interactions API for text/image-to-video and supports URI delivery for large videos.",
        "Context video provider audit/reference.",
    ),
]

DOWNLOAD_ASSETS: list[DownloadAsset] = [
    DownloadAsset(
        key="invention_jpg",
        url="https://commons.wikimedia.org/wiki/Special:Redirect/file/Angelica_Kauffmann_(1741-1807)_-_Invention_-_03-1128_-_Royal_Academy_of_Arts.jpg",
        dest="assets/public_domain/invention_kauffmann.jpg",
        category="visual",
        required=False,
        notes="Angelica Kauffmann 'Invention' reference image for clip 8.",
    ),
    DownloadAsset(
        key="osi_logo_png",
        url="https://i0.wp.com/opensource.org/wp-content/uploads/2023/03/cropped-OSI-horizontal-large.png?fit=640%2C229&quality=80&ssl=1",
        dest="assets/logos/osi_open_source_initiative.png",
        category="logo",
        required=False,
        notes="Official Open Source Initiative image from opensource.org. Preserve colors/aspect ratio.",
    ),
    DownloadAsset(
        key="oshw_logo_png",
        url="https://resources.oshwa.org/files/assets/oshw-logo-filled-color.png",
        dest="assets/logos/oshw_logo_filled_color.png",
        category="logo",
        required=False,
        notes="Open Source Hardware logo download from OSHWA resources. Preserve colors/aspect ratio.",
    ),
    DownloadAsset(
        key="beagleboard_logos_zip",
        url="https://www.beagleboard.org/app/uploads/downloads/beagleboard-logos.zip",
        dest="assets/logos/beagleboard-logos.zip",
        category="logo_zip",
        required=False,
        notes="Official BeagleBoard.org logo ZIP from brand-use page. Script extracts first usable PNG/SVG for end card.",
    ),
]

# -----------------------------------------------------------------------------
# Scenes
# -----------------------------------------------------------------------------

SCENES: list[Scene] = [
    Scene(
        1,
        "cold_open_second_question",
        20,
        "The second question in technology is the one we ask after the wreckage: why did it fail? Challenger. Chernobyl. Fukushima.",
        "LOCKED LOCAL RENDER, NOT SORA BY DEFAULT. Black screen. Centered phrase WHY DID IT FAIL appears, then cut into abstract technical mechanism visuals: pressure-seal cross-section, feedback-loop diagram, flooded backup-power schematic. Mechanism visuals only, no disaster spectacle.",
        ["WHY DID IT FAIL?"],
        ["challenger", "chernobyl", "fukushima"],
    ),
    Scene(
        2,
        "catastrophic_mechanisms_one",
        20,
        "At Challenger, pressure escaped where the vehicle could not stay whole. At Chernobyl, cooling conditions fed back into reactor power. At Fukushima, shutdown worked, but power for cooling was trapped behind one flooded assumption.",
        "Clean animated technical diagrams. First, right solid rocket motor aft field joint pressure-seal cross-section: pressure boundary, hot gas leak, vehicle integrity path breaks. Second, reactor feedback loop: steam bubbles increase reactivity, heat increases steam. Third, coastal nuclear power chain: earthquake shutdown, grid power lost, diesel generators, tsunami water flooding backup systems, batteries and cooling fading.",
        [
            "PRESSURE SEAL FAILURE -> HOT GAS LEAK -> NO ALTERNATE VEHICLE INTEGRITY",
            "ONE FEEDBACK LOOP",
            "COOLING CONDITION COUPLED TO REACTOR POWER",
            "ONE FLOOD",
            "PRIMARY AND BACKUP POWER EXPOSED TO SAME HAZARD",
        ],
        ["challenger", "chernobyl", "fukushima"],
    ),
    Scene(
        3,
        "catastrophic_mechanisms_two",
        20,
        "On the MAX, one angle-of-attack vane could feed repeated automated trim. On Deepwater Horizon, the last barrier could not seal. In the blackout, one disturbance became a regionwide cascade.",
        "Editorial technical animations. A single angle-of-attack vane sends data into an automated flight-control box; repeated nose-down trim arrows. Deepwater oil well cutaway: wellbore, blowout preventer, blind shear ram, buckled off-center pipe, and failure to seal. Grid map: one line trip shifts load to neighboring lines; more lines trip; map dims in cascade.",
        [
            "ONE SENSOR",
            "SINGLE SENSOR -> AUTOMATED AUTHORITY -> REPEATED COMMAND",
            "ONE FINAL BARRIER",
            "LAST BARRIER FAILED -> NO FAST CONTAINMENT PATH",
            "ONE GRID CASCADE",
            "DISTURBANCE NOT CONTAINED -> CASCADE",
        ],
        ["737max", "deepwater", "blackout2003"],
    ),
    Scene(
        4,
        "catastrophic_mechanisms_three",
        20,
        "Tacoma Narrows turned wind energy into self-excited deck motion. Hyatt Regency moved twice the load into a connection with no reserve. Boston stored a moving wall of molasses inside one failed vessel.",
        "Three restrained mechanism animations. Tacoma Narrows bridge as simplified flexible narrow deck cross-section: wind flow couples into twisting deck motion; do not show spectacular collapse. Hyatt Regency hanger rods: original continuous rod beside constructed interrupted rods; animate doubled load at fourth-floor box-beam connection. Molasses tank in dense city blocks: single storage vessel fails; streets show absence of secondary containment.",
        [
            "WIND ENERGY COUPLED INTO FLEXIBLE DECK MOTION",
            "ONE CONNECTION",
            "CONNECTION OVERLOAD -> NO RESERVE CAPACITY -> PROGRESSIVE COLLAPSE",
            "STORED HAZARD + FAILED VESSEL + NO CONTAINMENT",
        ],
        ["tacoma", "hyatt", "molasses"],
    ),
    Scene(
        5,
        "agency_and_alternatives",
        20,
        "And sometimes no one dies, so we stop calling it disaster: the Wi-Fi goes out and the house goes dumb. A lawsuit alleges a delivery route has no bathroom slack. A refrigerator becomes an ad screen. A paid-for smart cooker becomes a subscription. A cloud shutdown turns switches into paperweights. An app becomes the boss.",
        "Fast but calm sequence of everyday dependency failures. Smart home at night: Wi-Fi icon drops, lights and thermostat become unresponsive. Delivery route map with 200+ stops, countdown timers, bathroom icon off-route, surveillance dots. Smart refrigerator screen changes from family calendar to ad panel. Smart cooker diagram: paid device, remote server features, subscription gate, manual-only fallback. Cloud smart-home switches depend on vendor cloud; cloud disappears. Rideshare driver phone shows abstract ratings, GPS tracking, dispatch, dynamic pricing, support bot, deactivation risk. Include a half-second visual metaphor only: a giant carnival teddy bear behind one narrow path, no narration, no technological-failure label.",
        [
            "HOUSE CONTROL COUPLED TO WI-FI",
            "LAWSUIT ALLEGES: WORK PACE + TRACKING PREVENTED BREAKS",
            "APPLIANCE COUPLED TO VENDOR SOFTWARE POLICY",
            "DEVICE INTELLIGENCE LIVES SOMEWHERE ELSE",
            "HOME CONTROL COUPLED TO VENDOR SURVIVAL",
            "WORK COUPLED TO OPAQUE PLATFORM RULES",
        ],
        ["amazon_lawsuit", "smart_fridge_ads", "mellow", "insteon", "uber_unsw"],
    ),
    Scene(
        6,
        "reveal_structure",
        20,
        "These look like different failures. They are not. In each case, something became too essential, and there was no meaningful alternative path. That is tight coupling: failure stops being local because there is no meaningful substitute.",
        "Transform prior examples into simplified dependency diagrams. Many different systems collapse into the same structure: need, single solution, failure point, cascading harm. Diagram tiles align into a single red dependency path. The red path breaks; downstream nodes go dark.",
        [
            "NEED -> SINGLE SOLUTION -> FAILURE POINT -> CASCADING HARM",
            "NO MEANINGFUL ALTERNATIVE",
            "TIGHT COUPLING",
            "FAILURE STOPS BEING LOCAL",
        ],
        ["challenger", "chernobyl", "fukushima", "737max", "deepwater", "blackout2003"],
    ),
    Scene(
        7,
        "experts_build_tight_coupling",
        20,
        "The trigger may be bad design, bad incentives, bad luck, bad maintenance, bad assumptions, or unexpected conditions. But collapse occurs when there is no bypass, reserve, repair path, manual override, interoperability, or exit. And here is the terrifying part: experts build this. We all do. We call it efficiency, integration, convenience, optimization. But when alternatives disappear, those words become dependence. So how do we avoid building tight coupling?",
        "Editorial montage of engineering drawings, control dashboards, clean industrial diagrams, smart-home marketing layouts, route-optimization screens, platform app abstractions, and generic non-branded seamless integration copy. Backup paths fade out one by one. Bypass, reserve, repair path, manual override, interoperability, and exit each disappear from the diagram. Finish on near-black frame.",
        [
            "EFFICIENCY",
            "INTEGRATION",
            "CONVENIENCE",
            "OPTIMIZATION",
            "DEPENDENCE",
            "HOW DO WE AVOID IT?",
        ],
        ["ddex_ern"],
    ),
    Scene(
        8,
        "first_question_reveal",
        20,
        "Only now do we return to the first question of technology: what need are we actually trying to satisfy? Not the product. Not the platform. The need. Power. Communication. Transportation. Heat. Work. Safety.",
        "Use the input reference image of Angelica Kauffmann's Invention as a still-image visual anchor. Slow, respectful push-in. Do not add cliché phrases about invention. Let the classical image carry the idea. Then dissolve from the painting into a clean diagram: PRODUCT fades away, PLATFORM fades away, NEED remains. Needs appear as six columns: Power, Communication, Transportation, Heat, Work, Safety. Branching alternatives begin.",
        [
            "WHAT NEED ARE WE ACTUALLY TRYING TO SATISFY?",
            "NOT THE PRODUCT",
            "NOT THE PLATFORM",
            "THE NEED",
            "POWER  COMMUNICATION  TRANSPORTATION  HEAT  WORK  SAFETY",
        ],
        ["invention_image"],
        reference_asset_key="invention_jpg",
    ),
    Scene(
        9,
        "alternatives_network",
        16,
        "Power may be grid, battery, generator, solar, or local fallback. Communication may be internet, phone, mesh, radio, or a place to meet. Transportation may be walking, bike, transit, car, rideshare, or delivery. Heat may be utility service, backup heat, insulation, or passive design. Work may be an employer, clients, cooperative platforms, local networks, and tools owned by the worker. Safety may be sensors, procedures, human override, independent checks, and fail-safe design. A need is not reliably solved by one brittle answer. It is solved by a system of alternatives: backups, substitutes, repair, local control, open standards, interoperability, manual overrides, and exits. The trigger of failure changes. The structural source is the same: no meaningful alternative. If the need matters, the solution must survive the failure of any single part. If the need matters, build alternatives.",
        "A single red dependency path becomes a resilient network. One node fails; the system reroutes through other nodes. Branching alternatives appear as icon-like generic nodes: grid, battery, generator, solar, local fallback; internet, phone, mesh, radio, physical meeting; walking, bike, transit, car, rideshare, delivery; utility heat, backup heater, insulation, passive design; employer, clients, cooperative platforms, local networks, worker-owned tools; sensors, procedures, human override, independent checks, fail-safe design. End with blank placeholder area for official logos to be composited later; do not generate logos.",
        [
            "POWER: GRID | BATTERY | GENERATOR | SOLAR | LOCAL FALLBACK",
            "COMMUNICATION: INTERNET | PHONE | MESH | RADIO | MEETING",
            "TRANSPORTATION: WALKING | BIKE | TRANSIT | CAR | RIDESHARE | DELIVERY",
            "HEAT: UTILITY | BACKUP HEATER | INSULATION | PASSIVE DESIGN",
            "WORK: EMPLOYER | CLIENTS | COOPERATIVES | LOCAL NETWORKS | OWNED TOOLS",
            "SAFETY: SENSORS | PROCEDURES | HUMAN OVERRIDE | CHECKS | FAIL-SAFE DESIGN",
            "IF THE NEED MATTERS, BUILD ALTERNATIVES",
        ],
        ["osi_brand", "oshwa_logo", "beagle_brand"],
    ),
]

# -----------------------------------------------------------------------------
# Generic helpers
# -----------------------------------------------------------------------------

def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def read_size(size: str) -> tuple[int, int]:
    m = re.match(r"^(\d+)x(\d+)$", size)
    if not m:
        raise ValueError(f"Invalid size {size!r}; expected WIDTHxHEIGHT, for example 1920x1080")
    return int(m.group(1)), int(m.group(2))


def run_cmd(cmd: list[str], cwd: Optional[Path] = None) -> None:
    log("$ " + " ".join(cmd))
    subprocess.run(cmd, cwd=str(cwd) if cwd else None, check=True)


def check_cmd(name: str) -> None:
    if shutil.which(name) is None:
        raise RuntimeError(f"Missing required command on PATH: {name}")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def seconds_to_ass_time(seconds: float) -> str:
    if seconds < 0:
        seconds = 0
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h}:{m:02d}:{s:05.2f}"


def ffprobe_duration(path: Path) -> float:
    check_cmd("ffprobe")
    result = subprocess.run(
        [
            "ffprobe",
            "-v",
            "error",
            "-show_entries",
            "format=duration",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            str(path),
        ],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=True,
    )
    try:
        return float(result.stdout.strip())
    except ValueError:
        return 0.0


def find_font() -> Optional[Path]:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for c in candidates:
        p = Path(c)
        if p.exists():
            return p
    return None


def pil_font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    candidates = []
    if bold:
        candidates += [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
        ]
    candidates += [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for c in candidates:
        p = Path(c)
        if p.exists():
            return ImageFont.truetype(str(p), size=size)
    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    cur = ""
    for w in words:
        candidate = (cur + " " + w).strip()
        bbox = draw.textbbox((0, 0), candidate, font=font)
        if bbox[2] - bbox[0] <= max_width or not cur:
            cur = candidate
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def draw_wrapped(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: tuple[int, int, int] | str,
    max_width: int,
    line_gap: int = 8,
) -> int:
    x, y = xy
    for line in wrap_text(draw, text, font, max_width):
        draw.text((x, y), line, font=font, fill=fill)
        bbox = draw.textbbox((x, y), line, font=font)
        y += (bbox[3] - bbox[1]) + line_gap
    return y

# -----------------------------------------------------------------------------
# Asset downloads and source manifest
# -----------------------------------------------------------------------------

def download_url(url: str, dest: Path, required: bool = False, timeout: int = 60) -> bool:
    ensure_dir(dest.parent)
    if dest.exists() and dest.stat().st_size > 0:
        log(f"Asset exists: {dest}")
        return True
    headers = {"User-Agent": "second-question-production/1.0 (+source-audit)"}
    try:
        with requests.get(url, stream=True, headers=headers, timeout=timeout, allow_redirects=True) as r:
            r.raise_for_status()
            with dest.open("wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 256):
                    if chunk:
                        f.write(chunk)
        log(f"Downloaded: {dest}")
        return True
    except Exception as e:
        msg = f"Download failed for {url}: {e}"
        if required:
            raise RuntimeError(msg) from e
        log("WARNING: " + msg)
        return False


def write_sources_csv(root: Path) -> None:
    sources_dir = root / "sources"
    ensure_dir(sources_dir)
    out = sources_dir / "sources_manifest.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["key", "title", "url", "claim", "use_in_video", "license_note"])
        w.writeheader()
        for s in SOURCES:
            w.writerow(asdict(s))
    log(f"Wrote {out}")


def download_source_pages(root: Path) -> None:
    audit_dir = root / "sources" / "downloaded_pages"
    ensure_dir(audit_dir)
    index: list[dict[str, str]] = []
    for s in SOURCES:
        safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", s.key)[:80]
        dest = audit_dir / f"{safe}.html"
        ok = download_url(s.url, dest, required=False, timeout=60)
        index.append({"key": s.key, "url": s.url, "downloaded": str(dest if ok else ""), "ok": str(ok)})
    with (audit_dir / "download_index.json").open("w", encoding="utf-8") as f:
        json.dump(index, f, indent=2)


def extract_beagle_logo(root: Path) -> Optional[Path]:
    zip_path = root / "assets/logos/beagleboard-logos.zip"
    extract_dir = root / "assets/logos/beagleboard_extracted"
    if not zip_path.exists():
        return None
    ensure_dir(extract_dir)
    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(extract_dir)
    except Exception as e:
        log(f"WARNING: could not extract BeagleBoard ZIP: {e}")
        return None

    candidates = []
    for p in extract_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}:
            name = p.name.lower()
            score = 0
            if "beagleboard" in name:
                score += 10
            if "logo" in name:
                score += 5
            if "approved" in name or "compatible" in name or "embedded" in name:
                score -= 20
            candidates.append((score, p))
    if not candidates:
        log("WARNING: no PNG/JPEG logo found in BeagleBoard ZIP. Place official logo manually in assets/logos/beagleboard.png")
        return None
    candidates.sort(key=lambda x: (-x[0], len(str(x[1]))))
    chosen = candidates[0][1]
    dest = root / "assets/logos/beagleboard_official.png"
    shutil.copyfile(chosen, dest)
    log(f"Extracted BeagleBoard logo candidate: {dest} from {chosen}")
    return dest


def create_sora_reference_images(root: Path, size: str) -> None:
    w, h = read_size(size)
    src = root / "assets/public_domain/invention_kauffmann.jpg"
    if not src.exists():
        log("WARNING: Invention image missing; locked scene 8 will use a neutral placeholder unless you add the asset.")
        return
    out = root / "assets/reference_frames" / f"invention_ref_{w}x{h}.jpg"
    ensure_dir(out.parent)
    img = Image.open(src).convert("RGB")
    # Preserve the image; create a resolution-matched canvas because image reference
    # inputs are expected to match target video resolution.
    canvas = Image.new("RGB", (w, h), (8, 8, 10))
    img.thumbnail((int(w * 0.86), int(h * 0.86)), Image.LANCZOS)
    x = (w - img.width) // 2
    y = (h - img.height) // 2
    canvas.paste(img, (x, y))
    canvas.save(out, quality=95)
    log(f"Created Sora reference frame: {out}")


def command_download_assets(args: argparse.Namespace) -> None:
    root = Path(args.root)
    ensure_dir(root)
    write_sources_csv(root)
    if not args.no_source_pages:
        download_source_pages(root)
    asset_rows = []
    for asset in DOWNLOAD_ASSETS:
        dest = root / asset.dest
        ok = download_url(asset.url, dest, required=asset.required)
        asset_rows.append({**asdict(asset), "downloaded": ok, "sha256": sha256_file(dest) if dest.exists() else ""})
    extract_beagle_logo(root)
    create_sora_reference_images(root, args.size)
    manifest_path = root / "assets" / "asset_manifest.json"
    ensure_dir(manifest_path.parent)
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(asset_rows, f, indent=2)
    log(f"Wrote {manifest_path}")

# -----------------------------------------------------------------------------
# Context video generation and mock clips
# -----------------------------------------------------------------------------

def profile_for(key: str) -> dict[str, Any]:
    try:
        return CONTEXT_GENERATION_PROFILES[key]
    except KeyError as e:
        choices = ", ".join(sorted(CONTEXT_GENERATION_PROFILES))
        raise RuntimeError(f"Unknown context generation profile {key!r}. Choose in code from: {choices}") from e


def provider_output_name(provider_key: str) -> str:
    return "omni" if provider_key == "gemini_omni" else provider_key


def context_clips_dir(root: Path, provider_key: Optional[str]) -> Path:
    if provider_key is None:
        return root / "clips" / "mock"
    return root / "clips" / provider_output_name(provider_key)


def scene_is_locked_local(scene: Scene) -> bool:
    return scene.id in LOCKED_LOCAL_SCENE_IDS


def context_visual_prompt(scene: Scene, provider_key: str) -> str:
    source_notes = []
    src_by_key = {s.key: s for s in SOURCES}
    for key in scene.source_keys:
        s = src_by_key.get(key)
        if s:
            source_notes.append(f"- {s.claim}")
    notes = "\n".join(source_notes)
    labels = "\n".join([f"- {x}" for x in scene.labels])
    provider_hint = "Gemini Omni Flash" if provider_key == "gemini_omni" else "Sora"
    return f"""{MASTER_SORA_STYLE}

Provider context: {provider_hint}. This clip is background/context footage only. The final script, timing, labels, logos, cited images, and voiceover are locked in post-production.

Scene {scene.id}: {scene.slug.replace('_', ' ')}
Target scene duration in final edit: {scene.seconds} seconds. The generated clip can be shorter; it will be looped or trimmed deterministically in assembly.

Visual direction:
{scene.prompt}

Source-grounding notes for visual mechanism accuracy:
{notes}

Exact labels that will be composited in post, not rendered by the model:
{labels}

Additional constraints: no on-screen words, no captions, no lower thirds, no logos, no trademarks, no brand UI, no narration, no dialogue, no generated title cards, no disaster spectacle.
""".strip()


def sora_prompt(scene: Scene) -> str:
    return context_visual_prompt(scene, "sora")


def gemini_omni_prompt(scene: Scene) -> str:
    return context_visual_prompt(scene, "gemini_omni")


def requested_seconds_for(profile: dict[str, Any], scene: Scene) -> int:
    # AI context footage is intentionally decoupled from final timing.
    # Keep request length short for iteration/cost; assembly conforms duration.
    val = profile.get("request_seconds", min(scene.seconds, 8))
    try:
        return max(1, int(val))
    except Exception:
        return min(scene.seconds, 8)


def write_context_requests(root: Path, provider_key: str, size: str, scenes: Optional[list[Scene]] = None) -> None:
    profile = profile_for(provider_key)
    provider_dir = root / provider_output_name(provider_key)
    out = provider_dir / f"{provider_output_name(provider_key)}_requests.jsonl"
    ensure_dir(out.parent)
    scenes = scenes or SCENES
    with out.open("w", encoding="utf-8") as f:
        for scene in scenes:
            if scene_is_locked_local(scene):
                body = {"local_render": True, "reason": "locked local scene; no AI call"}
            elif provider_key == "sora":
                body = {
                    "model": profile["model"],
                    "prompt": context_visual_prompt(scene, provider_key),
                    "size": profile.get("size", size),
                    "seconds": str(requested_seconds_for(profile, scene)),
                }
                ref = local_reference_for_scene(root, scene, size)
                if ref:
                    body["input_reference"] = str(ref)
            elif provider_key == "gemini_omni":
                body = gemini_omni_payload(scene, profile, root, include_binary=False)
            else:
                body = {"error": f"Unsupported provider {provider_key}"}
            f.write(json.dumps({"scene_id": scene.id, "custom_id": f"scene-{scene.id:02d}", "provider": provider_key, "body": body}) + "\n")
    log(f"Wrote dry-run {provider_key} request plan: {out}")


def write_sora_requests(root: Path, model: str, size: str) -> None:
    # Compatibility wrapper for older scripts; model selection now lives in CONTEXT_GENERATION_PROFILES.
    write_context_requests(root, "sora", size)


def local_reference_for_scene(root: Path, scene: Scene, size: str) -> Optional[Path]:
    # Specific images are rendered/composited locally in the final pipeline. This
    # function remains for experiments, but locked local scenes should not call AI.
    if scene.reference_asset_key == "invention_jpg":
        w, h = read_size(size)
        p = root / "assets/reference_frames" / f"invention_ref_{w}x{h}.jpg"
        if p.exists():
            return p
    return None


def openai_headers() -> dict[str, str]:
    key = os.getenv("OPENAI_API_KEY")
    if not key:
        raise RuntimeError("OPENAI_API_KEY is not set.")
    return {"Authorization": f"Bearer {key}"}


def gemini_key() -> str:
    key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not key:
        raise RuntimeError("GEMINI_API_KEY is not set. GOOGLE_API_KEY is also accepted.")
    return key


def file_to_data_url(path: Path) -> str:
    """Return a base64 data URL usable as Videos API input_reference.image_url."""
    mime = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def create_sora_job(scene: Scene, root: Path, profile: dict[str, Any], final_size: str) -> str:
    url = f"{API_BASE}/videos"
    payload: dict[str, Any] = {
        "model": profile["model"],
        "prompt": context_visual_prompt(scene, "sora"),
        "size": profile.get("size", final_size),
        "seconds": str(requested_seconds_for(profile, scene)),
    }
    ref = local_reference_for_scene(root, scene, final_size)
    if ref:
        # The Videos API create endpoint accepts JSON. For image guidance,
        # input_reference should be an object with either image_url or file_id.
        # A local reference image can be sent as a base64 data URL.
        payload["input_reference"] = {"image_url": file_to_data_url(ref)}

    r = requests.post(url, headers=openai_headers(), json=payload, timeout=120)
    if r.status_code >= 400:
        raise RuntimeError(f"Sora create failed for scene {scene.id}: HTTP {r.status_code}\n{r.text}")
    obj = r.json()
    job_id = obj.get("id")
    if not job_id:
        raise RuntimeError(f"Sora create response missing id: {obj}")
    return job_id


def poll_sora_job(job_id: str, poll_seconds: int = 15, timeout_minutes: int = 90) -> dict[str, Any]:
    deadline = time.time() + timeout_minutes * 60
    url = f"{API_BASE}/videos/{job_id}"
    last_status = None
    while True:
        r = requests.get(url, headers=openai_headers(), timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"Sora retrieve failed for {job_id}: HTTP {r.status_code}\n{r.text}")
        obj = r.json()
        status = obj.get("status")
        progress = obj.get("progress")
        if status != last_status:
            log(f"Sora job {job_id}: status={status} progress={progress}")
            last_status = status
        if status == "completed":
            return obj
        if status == "failed":
            raise RuntimeError(f"Sora job failed {job_id}: {obj.get('error')}")
        if time.time() > deadline:
            raise TimeoutError(f"Timed out waiting for Sora job {job_id}")
        time.sleep(poll_seconds)


def download_sora_content(job_id: str, dest: Path) -> None:
    ensure_dir(dest.parent)
    url = f"{API_BASE}/videos/{job_id}/content"
    with requests.get(url, headers=openai_headers(), params={"variant": "video"}, stream=True, timeout=300) as r:
        if r.status_code >= 400:
            raise RuntimeError(f"Sora content download failed for {job_id}: HTTP {r.status_code}\n{r.text}")
        with dest.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)
    log(f"Downloaded Sora context clip: {dest}")


def gemini_omni_payload(scene: Scene, profile: dict[str, Any], root: Path, include_binary: bool = True) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "model": profile["model"],
        "input": gemini_omni_prompt(scene),
        "response_format": {
            "type": "video",
            "aspect_ratio": profile.get("aspect_ratio", "16:9"),
            "delivery": profile.get("delivery", "uri"),
        },
        # Fast synchronous context generation. Leave store=True in the profile if
        # you want to use previous_interaction_id for conversational edits.
        "background": bool(profile.get("background", False)),
        "store": bool(profile.get("store", False)),
        "stream": bool(profile.get("stream", False)),
    }
    task = profile.get("task")
    if task:
        payload["generation_config"] = {"video_config": {"task": task}}
    if not include_binary:
        return payload
    return payload


def extract_gemini_video_object(obj: dict[str, Any]) -> dict[str, Any]:
    direct = obj.get("output_video")
    if isinstance(direct, dict):
        return direct
    for step in obj.get("steps", []) or []:
        if not isinstance(step, dict):
            continue
        for item in step.get("content", []) or []:
            if isinstance(item, dict) and item.get("type") == "video":
                return item
    raise RuntimeError(f"Gemini Omni response did not include a video object: {json.dumps(obj)[:2000]}")


def download_gemini_uri(uri: str, dest: Path, poll_seconds: int, timeout_minutes: int) -> None:
    ensure_dir(dest.parent)
    key = gemini_key()
    base = os.getenv("GEMINI_API_BASE", "https://generativelanguage.googleapis.com/v1beta")
    file_id_match = re.search(r"files/([^/:?]+)", uri)
    if file_id_match:
        file_id = file_id_match.group(1)
        deadline = time.time() + timeout_minutes * 60
        last_state = None
        while True:
            r = requests.get(f"{base}/files/{file_id}", params={"key": key}, timeout=60)
            if r.status_code >= 400:
                raise RuntimeError(f"Gemini file status failed for {file_id}: HTTP {r.status_code}\n{r.text}")
            status_obj = r.json()
            state = status_obj.get("state")
            if isinstance(state, dict):
                state = state.get("name")
            if state != last_state:
                log(f"Gemini file {file_id}: state={state}")
                last_state = state
            if state in ("ACTIVE", "SUCCEEDED", "READY"):
                break
            if state in ("FAILED", "ERROR"):
                raise RuntimeError(f"Gemini file failed: {status_obj}")
            if time.time() > deadline:
                raise TimeoutError(f"Timed out waiting for Gemini file {file_id}")
            time.sleep(poll_seconds)
        download_url = f"{base}/files/{file_id}:download"
        with requests.get(download_url, params={"alt": "media", "key": key}, stream=True, timeout=300) as r:
            if r.status_code >= 400:
                raise RuntimeError(f"Gemini file download failed for {file_id}: HTTP {r.status_code}\n{r.text}")
            with dest.open("wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 1024):
                    if chunk:
                        f.write(chunk)
        return
    # Fallback: try the URI directly. Some responses provide a complete download URI.
    with requests.get(uri, params={"key": key}, stream=True, timeout=300) as r:
        if r.status_code >= 400:
            raise RuntimeError(f"Gemini URI download failed: HTTP {r.status_code}\n{r.text}")
        with dest.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)


def create_gemini_omni_clip(scene: Scene, root: Path, profile: dict[str, Any], dest: Path, poll_seconds: int, timeout_minutes: int) -> dict[str, Any]:
    base = os.getenv("GEMINI_API_BASE", "https://generativelanguage.googleapis.com/v1beta")
    url = f"{base}/interactions"
    headers = {"x-goog-api-key": gemini_key(), "Content-Type": "application/json"}
    payload = gemini_omni_payload(scene, profile, root)
    log(f"Starting Gemini Omni context clip for scene {scene.id}: {scene.slug}")
    r = requests.post(url, headers=headers, json=payload, timeout=300)
    if r.status_code >= 400:
        raise RuntimeError(f"Gemini Omni create failed for scene {scene.id}: HTTP {r.status_code}\n{r.text}")
    obj = r.json()
    log_dir = root / "gemini_omni"
    ensure_dir(log_dir)
    (log_dir / f"scene_{scene.id:02d}_interaction.json").write_text(json.dumps(obj, indent=2), encoding="utf-8")
    video = extract_gemini_video_object(obj)
    ensure_dir(dest.parent)
    if video.get("data"):
        dest.write_bytes(base64.b64decode(video["data"]))
    elif video.get("uri"):
        download_gemini_uri(video["uri"], dest, poll_seconds, timeout_minutes)
    else:
        raise RuntimeError(f"Gemini video object had neither data nor uri: {video}")
    log(f"Downloaded Gemini Omni context clip: {dest}")
    return obj


def render_locked_local_scene(scene: Scene, root: Path, out: Path, size: str, fps: int = DEFAULT_FPS) -> None:
    if scene.id == 1:
        render_locked_cold_open(scene, out, size, fps=fps)
    elif scene.id == 8:
        render_locked_invention_reveal(scene, root, out, size, fps=fps)
    elif scene.id == 9:
        render_locked_alternatives_network(scene, root, out, size, fps=fps)
    else:
        render_mock_clip(scene, out, size, fps=fps)


def generate_ai_context_clip(provider_key: str, scene: Scene, root: Path, out: Path, size: str, poll_seconds: int, timeout_minutes: int, force: bool = False, fps: int = DEFAULT_FPS) -> None:
    profile = profile_for(provider_key)
    if scene_is_locked_local(scene):
        if out.exists() and not force:
            log(f"Locked local clip exists, skipping: {out}")
            return
        render_locked_local_scene(scene, root, out, size, fps=fps)
        return
    if out.exists() and not force:
        log(f"Context clip exists, skipping: {out}")
        return
    if provider_key == "sora":
        jobs_path = root / "sora" / "jobs.json"
        jobs: dict[str, Any] = {}
        if jobs_path.exists() and not force:
            jobs = json.loads(jobs_path.read_text(encoding="utf-8"))
        jid = jobs.get(str(scene.id), {}).get("job_id")
        if not jid or force:
            log(f"Starting Sora context job for scene {scene.id}: {scene.slug}")
            jid = create_sora_job(scene, root, profile, size)
            jobs[str(scene.id)] = {"job_id": jid, "scene": scene.slug, "created_at": datetime.now(timezone.utc).isoformat(), "profile": provider_key, "model": profile.get("model")}
            ensure_dir(jobs_path.parent)
            jobs_path.write_text(json.dumps(jobs, indent=2), encoding="utf-8")
        poll_sora_job(jid, poll_seconds=poll_seconds, timeout_minutes=timeout_minutes)
        download_sora_content(jid, out)
    elif provider_key == "gemini_omni":
        create_gemini_omni_clip(scene, root, profile, out, poll_seconds=poll_seconds, timeout_minutes=timeout_minutes)
    else:
        raise RuntimeError(f"Unsupported context provider: {provider_key}")


def command_generate_clips(args: argparse.Namespace) -> None:
    root = Path(args.root)
    provider_key: Optional[str]
    if getattr(args, "sora", False) and getattr(args, "omni", False):
        raise RuntimeError("Choose only one provider flag: --sora or --omni.")
    if getattr(args, "sora", False):
        provider_key = "sora"
    elif getattr(args, "omni", False):
        provider_key = "gemini_omni"
    else:
        provider_key = None
    clips_dir = context_clips_dir(root, provider_key)
    ensure_dir(clips_dir)
    only_scene = getattr(args, "only_scene", None)
    scenes = [scene for scene in SCENES if only_scene in (None, 0) or scene.id == only_scene]
    if not scenes:
        raise RuntimeError(f"No scene found for --only-scene {only_scene}")
    if args.dry_run:
        if provider_key is None:
            log("Dry run for mock/local clips: no API requests to write.")
        else:
            write_context_requests(root, provider_key, args.size, scenes=scenes)
        return
    if provider_key:
        create_sora_reference_images(root, args.size)
        for scene in scenes:
            out = clips_dir / scene.filename
            generate_ai_context_clip(provider_key, scene, root, out, args.size, args.poll_seconds, args.timeout_minutes, force=args.force, fps=args.fps)
    else:
        for scene in scenes:
            out = clips_dir / scene.filename
            if scene_is_locked_local(scene):
                if out.exists() and not args.force:
                    log(f"Locked local clip exists, skipping: {out}")
                else:
                    render_locked_local_scene(scene, root, out, args.size, fps=args.fps)
            elif out.exists() and not args.force:
                log(f"Mock clip exists, skipping: {out}")
            else:
                render_mock_clip(scene, out, args.size, fps=args.fps)


def command_compare_context_generators(args: argparse.Namespace) -> None:
    """Generate the same configured test scenes with both AI providers.

    Provider/model choices are controlled by COMPARE_CONTEXT_PROFILES and
    CONTEXT_GENERATION_PROFILES near the top of the file, not by CLI options.
    """
    root = Path(args.root)
    create_sora_reference_images(root, args.size)
    scenes = [scene for scene in SCENES if scene.id in COMPARE_CONTEXT_SCENE_IDS]
    if args.dry_run:
        for provider_key in COMPARE_CONTEXT_PROFILES:
            write_context_requests(root, provider_key, args.size, scenes=scenes)
        return
    for provider_key in COMPARE_CONTEXT_PROFILES:
        compare_dir = root / "clips" / "compare" / provider_output_name(provider_key)
        ensure_dir(compare_dir)
        for scene in scenes:
            out = compare_dir / scene.filename
            generate_ai_context_clip(provider_key, scene, root, out, args.size, args.poll_seconds, args.timeout_minutes, force=args.force, fps=args.fps)
    write_compare_readme(root, scenes)


def write_compare_readme(root: Path, scenes: list[Scene]) -> None:
    out = root / "clips" / "compare" / "README_compare_context_generators.md"
    ensure_dir(out.parent)
    lines = [
        "# Context generator comparison\n",
        "Generated the same configured test scenes with each provider. The final cut should still strip generated audio, normalize duration, and burn labels/logos/images locally.\n",
        "## Scenes\n",
    ]
    for scene in scenes:
        lines.append(f"- Scene {scene.id}: `{scene.slug}` ({scene.seconds}s target)\n")
    lines.append("\n## Providers\n")
    for key in COMPARE_CONTEXT_PROFILES:
        profile = profile_for(key)
        lines.append(f"- `{key}`: model `{profile.get('model')}`; outputs in `clips/compare/{provider_output_name(key)}/`\n")
    out.write_text("".join(lines), encoding="utf-8")
    log(f"Wrote comparison notes: {out}")

def render_mock_clip(scene: Scene, out: Path, size: str, fps: int = DEFAULT_FPS) -> None:
    check_cmd("ffmpeg")
    w, h = read_size(size)
    ensure_dir(out.parent)
    frame = out.with_suffix(".png")
    img = Image.new("RGB", (w, h), (7, 7, 10))
    draw = ImageDraw.Draw(img)
    title_font = pil_font(max(24, w // 36), bold=True)
    body_font = pil_font(max(16, w // 64), bold=False)
    small_font = pil_font(max(14, w // 78), bold=False)
    red = (205, 60, 60)
    white = (245, 245, 245)
    grey = (150, 155, 160)
    dark_panel = (22, 24, 28)

    # Background grid
    step = max(80, w // 16)
    for x in range(0, w, step):
        draw.line((x, 0, x, h), fill=(16, 17, 20), width=1)
    for y in range(0, h, step):
        draw.line((0, y, w, y), fill=(16, 17, 20), width=1)

    # Scene header
    draw.text((int(w * 0.05), int(h * 0.06)), f"SCENE {scene.id:02d}", font=small_font, fill=grey)
    draw_wrapped(draw, (int(w * 0.05), int(h * 0.10)), scene.slug.replace("_", " ").upper(), title_font, white, int(w * 0.72))

    # Simple dependency diagram
    y = int(h * 0.42)
    x0 = int(w * 0.08)
    box_w = int(w * 0.20)
    box_h = int(h * 0.10)
    gap = int(w * 0.04)
    boxes = ["NEED", "SINGLE SOLUTION", "FAILURE POINT", "CASCADING HARM"]
    for i, b in enumerate(boxes):
        x = x0 + i * (box_w + gap)
        draw.rounded_rectangle((x, y, x + box_w, y + box_h), radius=12, fill=dark_panel, outline=grey, width=2)
        draw.text((x + 18, y + box_h // 2 - 12), b, font=small_font, fill=white)
        if i < len(boxes) - 1:
            ax1 = x + box_w
            ax2 = x + box_w + gap
            ay = y + box_h // 2
            draw.line((ax1, ay, ax2, ay), fill=red, width=5)
            draw.polygon([(ax2, ay), (ax2 - 14, ay - 9), (ax2 - 14, ay + 9)], fill=red)

    # Labels panel
    lx = int(w * 0.05)
    ly = int(h * 0.62)
    draw.text((lx, ly), "POST-PRODUCTION LABELS", font=small_font, fill=grey)
    ly += int(h * 0.045)
    for lab in scene.labels[:6]:
        draw_wrapped(draw, (lx, ly), lab, body_font, white, int(w * 0.86), line_gap=4)
        ly += int(h * 0.055)
        if ly > h - int(h * 0.08):
            break

    img.save(frame)
    run_cmd([
        "ffmpeg", "-y",
        "-loop", "1",
        "-i", str(frame),
        "-t", str(scene.seconds),
        "-r", str(fps),
        "-pix_fmt", "yuv420p",
        "-c:v", "libx264",
        "-movflags", "+faststart",
        str(out),
    ])
    log(f"Rendered mock clip: {out}")


def render_locked_cold_open(scene: Scene, out: Path, size: str, fps: int = DEFAULT_FPS) -> None:
    """Render scene 1 locally so the opening cannot drift from the script.

    Sora is useful for atmospheric mechanism visuals, but the cold open is a
    thesis-setting beat. This deterministic renderer keeps the black opening,
    exact title text, and three mechanism proof-point hints under editorial
    control. It produces a silent MP4; assembly later strips any source clip
    audio anyway and muxes the locked narration.
    """
    check_cmd("ffmpeg")
    w, h = read_size(size)
    ensure_dir(out.parent)
    work = out.parent / f".{out.stem}_locked_frames"
    ensure_dir(work)

    bg = (4, 4, 7)
    panel = (16, 18, 22)
    white = (245, 245, 245)
    grey = (142, 148, 156)
    dim = (60, 64, 72)
    red = (205, 60, 60)
    blue = (70, 125, 180)

    title_font = pil_font(max(56, w // 17), bold=True)
    tiny_font = pil_font(max(16, w // 96), bold=False)

    def base() -> Image.Image:
        img = Image.new("RGB", (w, h), bg)
        d = ImageDraw.Draw(img)
        step = max(90, w // 18)
        for x in range(0, w, step):
            d.line((x, 0, x, h), fill=(10, 11, 15), width=1)
        for y in range(0, h, step):
            d.line((0, y, w, y), fill=(10, 11, 15), width=1)
        return img

    def centered_text(draw: ImageDraw.ImageDraw, text: str, y: int, font: ImageFont.ImageFont, fill: tuple[int, int, int]) -> None:
        bbox = draw.textbbox((0, 0), text, font=font)
        draw.text(((w - (bbox[2] - bbox[0])) // 2, y), text, font=font, fill=fill)

    def arrow(draw: ImageDraw.ImageDraw, a: tuple[int, int], b: tuple[int, int], fill=red, width_px: int = 5) -> None:
        draw.line((a[0], a[1], b[0], b[1]), fill=fill, width=width_px)
        import math
        ang = math.atan2(b[1] - a[1], b[0] - a[0])
        size_arrow = max(10, w // 100)
        p1 = (b[0] - size_arrow * math.cos(ang - 0.45), b[1] - size_arrow * math.sin(ang - 0.45))
        p2 = (b[0] - size_arrow * math.cos(ang + 0.45), b[1] - size_arrow * math.sin(ang + 0.45))
        draw.polygon([b, p1, p2], fill=fill)

    def save_still(name: str, img: Image.Image) -> Path:
        path = work / name
        img.save(path)
        return path

    # 0.0-1.2: true black.
    img0 = Image.new("RGB", (w, h), bg)
    p0 = save_still("00_black.png", img0)

    # 1.2-6.8: locked title card.
    img1 = Image.new("RGB", (w, h), bg)
    d = ImageDraw.Draw(img1)
    centered_text(d, "WHY DID IT FAIL?", int(h * 0.43), title_font, white)
    d.line((int(w * 0.32), int(h * 0.55), int(w * 0.68), int(h * 0.55)), fill=red, width=max(3, w // 400))
    p1 = save_still("01_why.png", img1)

    # 6.8-11.2: pressure-seal cross-section visual hint.
    img2 = base(); d = ImageDraw.Draw(img2)
    cx, cy = int(w * 0.50), int(h * 0.52)
    # Rocket motor shell and joint
    d.rounded_rectangle((int(w*0.18), int(h*0.36), int(w*0.82), int(h*0.66)), radius=18, outline=grey, width=3, fill=panel)
    d.rectangle((int(w*0.46), int(h*0.31), int(w*0.54), int(h*0.71)), outline=white, width=4, fill=(24, 26, 31))
    d.ellipse((cx-int(w*0.035), cy-int(w*0.035), cx+int(w*0.035), cy+int(w*0.035)), outline=white, width=4)
    # Hot gas path and broken integrity route
    arrow(d, (int(w*0.28), cy), (cx-int(w*0.045), cy), fill=red, width_px=max(4, w//320))
    arrow(d, (cx+int(w*0.045), cy), (int(w*0.74), int(h*0.45)), fill=red, width_px=max(4, w//320))
    for x in [int(w*0.25), int(w*0.37), int(w*0.63), int(w*0.75)]:
        d.rectangle((x-28, int(h*0.25), x+28, int(h*0.29)), outline=dim, width=2)
        d.line((x, int(h*0.29), x, int(h*0.36)), fill=dim, width=2)
    p2 = save_still("02_seal.png", img2)

    # 11.2-15.6: feedback-loop visual hint.
    img3 = base(); d = ImageDraw.Draw(img3)
    points = [
        (int(w*0.50), int(h*0.22)),
        (int(w*0.72), int(h*0.45)),
        (int(w*0.60), int(h*0.72)),
        (int(w*0.35), int(h*0.68)),
        (int(w*0.28), int(h*0.40)),
    ]
    for i, a in enumerate(points):
        b = points[(i + 1) % len(points)]
        arrow(d, a, b, fill=red if i in (1,2) else grey, width_px=max(3, w//420))
    for i, (x, y) in enumerate(points):
        r = int(w * (0.032 if i == 0 else 0.026))
        d.ellipse((x-r, y-r, x+r, y+r), outline=white, width=3, fill=panel)
        # Blank callout boxes, labels added later in post-production.
        if i % 2 == 0:
            d.rounded_rectangle((x+int(w*0.035), y-int(h*0.025), x+int(w*0.16), y+int(h*0.025)), radius=6, outline=dim, width=2)
        else:
            d.rounded_rectangle((x-int(w*0.16), y-int(h*0.025), x-int(w*0.035), y+int(h*0.025)), radius=6, outline=dim, width=2)
    p3 = save_still("03_loop.png", img3)

    # 15.6-20.0: flooded backup-power schematic visual hint.
    img4 = base(); d = ImageDraw.Draw(img4)
    ground = int(h * 0.66)
    water = int(h * 0.72)
    d.rectangle((int(w*0.12), ground, int(w*0.88), int(h*0.82)), fill=(9, 15, 23), outline=dim)
    d.rectangle((int(w*0.12), water, int(w*0.88), int(h*0.82)), fill=(20, 60, 88), outline=None)
    # Plant and backup blocks
    d.rectangle((int(w*0.20), int(h*0.34), int(w*0.34), ground), outline=white, width=3, fill=panel)
    d.rectangle((int(w*0.45), int(h*0.45), int(w*0.56), ground), outline=grey, width=3, fill=(25, 28, 33))
    d.rectangle((int(w*0.62), int(h*0.45), int(w*0.73), ground), outline=grey, width=3, fill=(25, 28, 33))
    arrow(d, (int(w*0.34), int(h*0.52)), (int(w*0.45), int(h*0.52)), fill=grey, width_px=max(3, w//440))
    arrow(d, (int(w*0.56), int(h*0.52)), (int(w*0.62), int(h*0.52)), fill=grey, width_px=max(3, w//440))
    # Shared hazard cuts both primary and backup paths.
    d.line((int(w*0.12), water, int(w*0.88), water), fill=blue, width=max(5, w//300))
    for x in [int(w*0.50), int(w*0.67)]:
        d.line((x-int(w*0.035), int(h*0.44), x+int(w*0.035), int(h*0.63)), fill=red, width=max(5, w//300))
        d.line((x+int(w*0.035), int(h*0.44), x-int(w*0.035), int(h*0.63)), fill=red, width=max(5, w//300))
    p4 = save_still("04_flood.png", img4)

    # Use deterministic still durations. The final repeated file is required by
    # the concat demuxer for the last duration to take effect.
    concat = work / "concat.txt"
    schedule = [(p0, 1.2), (p1, 5.6), (p2, 4.4), (p3, 4.4), (p4, 4.4)]
    with concat.open("w", encoding="utf-8") as f:
        for path, dur in schedule:
            f.write(f"file '{path.resolve().as_posix()}'\n")
            f.write(f"duration {dur:.3f}\n")
        f.write(f"file '{p4.resolve().as_posix()}'\n")
    run_cmd([
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", str(concat),
        "-vf", f"fps={fps},format=yuv420p",
        "-t", f"{scene.seconds:.3f}",
        "-an",
        "-c:v", "libx264",
        "-preset", "medium",
        "-crf", "18",
        "-movflags", "+faststart",
        str(out),
    ])
    log(f"Rendered locked cold-open clip: {out}")


def render_still_schedule(schedule: list[tuple[Path, float]], out: Path, seconds: float, fps: int) -> None:
    """Render a deterministic silent video from a list of stills and durations."""
    check_cmd("ffmpeg")
    ensure_dir(out.parent)
    concat = out.parent / f".{out.stem}_concat.txt"
    with concat.open("w", encoding="utf-8") as f:
        for path, dur in schedule:
            f.write(f"file '{path.resolve().as_posix()}'\n")
            f.write(f"duration {dur:.3f}\n")
        # concat demuxer requires a final repeated file for the last duration.
        f.write(f"file '{schedule[-1][0].resolve().as_posix()}'\n")
    run_cmd([
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", str(concat),
        "-vf", f"fps={fps},format=yuv420p",
        "-t", f"{seconds:.3f}",
        "-an",
        "-c:v", "libx264",
        "-preset", "medium",
        "-crf", "18",
        "-movflags", "+faststart",
        str(out),
    ])


def render_locked_invention_reveal(scene: Scene, root: Path, out: Path, size: str, fps: int = DEFAULT_FPS) -> None:
    """Render the Kauffmann/first-question beat locally.

    This keeps the specific painting under source/asset control instead of asking
    an AI video model to reproduce it. Text remains out of the footage and is
    added through the ASS label pass.
    """
    w, h = read_size(size)
    ensure_dir(out.parent)
    work = out.parent / f".{out.stem}_locked_frames"
    ensure_dir(work)
    bg = (4, 4, 7)
    panel = (16, 18, 22)
    white = (245, 245, 245)
    grey = (142, 148, 156)
    dim = (60, 64, 72)
    red = (205, 60, 60)

    def base() -> Image.Image:
        img = Image.new("RGB", (w, h), bg)
        d = ImageDraw.Draw(img)
        step = max(90, w // 18)
        for x in range(0, w, step):
            d.line((x, 0, x, h), fill=(10, 11, 15), width=1)
        for y in range(0, h, step):
            d.line((0, y, w, y), fill=(10, 11, 15), width=1)
        return img

    def save(name: str, img: Image.Image) -> Path:
        path = work / name
        img.save(path)
        return path

    img1 = base(); d = ImageDraw.Draw(img1)
    art_path = root / "assets/public_domain/invention_kauffmann.jpg"
    if art_path.exists():
        art = Image.open(art_path).convert("RGB")
        art.thumbnail((int(w * 0.58), int(h * 0.72)), Image.LANCZOS)
        ax, ay = int(w * 0.10), int(h * 0.14)
        d.rounded_rectangle((ax - 18, ay - 18, ax + art.width + 18, ay + art.height + 18), radius=12, outline=dim, width=2, fill=panel)
        img1.paste(art, (ax, ay))
    else:
        ax, ay = int(w * 0.10), int(h * 0.14)
        d.rounded_rectangle((ax, ay, int(w * 0.58), int(h * 0.80)), radius=12, outline=dim, width=2, fill=panel)
        for i in range(8):
            d.arc((ax + i*18, ay + i*14, int(w*0.55)-i*18, int(h*0.78)-i*14), start=210, end=330, fill=grey, width=2)
    # Blank exact-question area; text composited later.
    d.rounded_rectangle((int(w*0.67), int(h*0.30), int(w*0.92), int(h*0.58)), radius=18, outline=dim, width=2, fill=(10, 12, 16))
    d.line((int(w*0.70), int(h*0.47), int(w*0.89), int(h*0.47)), fill=red, width=max(3, w//380))
    p1 = save("01_invention.png", img1)

    img2 = base(); d = ImageDraw.Draw(img2)
    # Product/platform fades away toward need: use boxes with no text.
    boxes = [
        (int(w*0.10), int(h*0.38), int(w*0.30), int(h*0.53)),
        (int(w*0.40), int(h*0.38), int(w*0.60), int(h*0.53)),
        (int(w*0.70), int(h*0.34), int(w*0.90), int(h*0.57)),
    ]
    for i, box in enumerate(boxes):
        d.rounded_rectangle(box, radius=14, outline=grey if i < 2 else white, width=3, fill=panel)
        if i < 2:
            d.line((box[0]+15, box[1]+15, box[2]-15, box[3]-15), fill=dim, width=3)
            d.line((box[2]-15, box[1]+15, box[0]+15, box[3]-15), fill=dim, width=3)
    for i in range(2):
        x1 = boxes[i][2]
        x2 = boxes[i+1][0]
        y = (boxes[i][1] + boxes[i][3]) // 2
        d.line((x1, y, x2, y), fill=red if i == 1 else grey, width=max(4, w//380))
        d.polygon([(x2, y), (x2-14, y-9), (x2-14, y+9)], fill=red if i == 1 else grey)
    p2 = save("02_need_not_product.png", img2)

    img3 = base(); d = ImageDraw.Draw(img3)
    center = (int(w*0.50), int(h*0.30))
    d.ellipse((center[0]-38, center[1]-38, center[0]+38, center[1]+38), outline=white, width=3, fill=panel)
    branch_y = int(h*0.62)
    xs = [int(w*(0.14 + i*0.145)) for i in range(6)]
    for x in xs:
        d.line((center[0], center[1]+38, x, branch_y-32), fill=grey, width=3)
        d.ellipse((x-32, branch_y-32, x+32, branch_y+32), outline=white, width=3, fill=panel)
        # Each need gets several small alternative nodes, unlabeled.
        for j in range(3):
            ox = x + (j-1)*int(w*0.035)
            oy = branch_y + int(h*0.12) + (j % 2)*int(h*0.04)
            d.line((x, branch_y+32, ox, oy-18), fill=dim, width=2)
            d.ellipse((ox-18, oy-18, ox+18, oy+18), outline=dim, width=2, fill=(10, 12, 16))
    p3 = save("03_needs_branch.png", img3)

    render_still_schedule([(p1, 7.0), (p2, 5.0), (p3, scene.seconds - 12.0)], out, scene.seconds, fps)
    log(f"Rendered locked invention/need reveal clip: {out}")


def render_locked_alternatives_network(scene: Scene, root: Path, out: Path, size: str, fps: int = DEFAULT_FPS) -> None:
    """Render the alternatives-network conclusion locally without labels/logos."""
    w, h = read_size(size)
    ensure_dir(out.parent)
    work = out.parent / f".{out.stem}_locked_frames"
    ensure_dir(work)
    bg = (4, 4, 7)
    panel = (16, 18, 22)
    white = (245, 245, 245)
    grey = (142, 148, 156)
    dim = (60, 64, 72)
    red = (205, 60, 60)
    green = (105, 180, 130)

    def base() -> Image.Image:
        img = Image.new("RGB", (w, h), bg)
        d = ImageDraw.Draw(img)
        step = max(90, w // 18)
        for x in range(0, w, step):
            d.line((x, 0, x, h), fill=(10, 11, 15), width=1)
        for y in range(0, h, step):
            d.line((0, y, w, y), fill=(10, 11, 15), width=1)
        return img

    def save(name: str, img: Image.Image) -> Path:
        path = work / name
        img.save(path)
        return path

    def arrow(d: ImageDraw.ImageDraw, a: tuple[int,int], b: tuple[int,int], color=red, width_px=5) -> None:
        d.line((a[0], a[1], b[0], b[1]), fill=color, width=width_px)
        import math
        ang = math.atan2(b[1]-a[1], b[0]-a[0])
        size_arrow = max(10, w // 110)
        p1 = (b[0] - size_arrow * math.cos(ang - 0.45), b[1] - size_arrow * math.sin(ang - 0.45))
        p2 = (b[0] - size_arrow * math.cos(ang + 0.45), b[1] - size_arrow * math.sin(ang + 0.45))
        d.polygon([b, p1, p2], fill=color)

    # Single brittle path.
    img1 = base(); d = ImageDraw.Draw(img1)
    y = int(h*0.50)
    xs = [int(w*x) for x in (0.16, 0.38, 0.60, 0.82)]
    for i, x in enumerate(xs):
        d.rounded_rectangle((x-70, y-42, x+70, y+42), radius=14, outline=white if i in (0,3) else grey, width=3, fill=panel)
        if i < len(xs)-1:
            arrow(d, (x+70, y), (xs[i+1]-70, y), red, max(5, w//360))
    p1 = save("01_single_path.png", img1)

    # Broken path.
    img2 = img1.copy(); d = ImageDraw.Draw(img2)
    bx = int((xs[1] + xs[2]) / 2)
    d.line((bx-35, y-48, bx+35, y+48), fill=red, width=max(8, w//240))
    d.line((bx+35, y-48, bx-35, y+48), fill=red, width=max(8, w//240))
    d.rectangle((xs[2]-70, y-42, xs[3]+80, y+45), outline=None, fill=(2, 2, 4))
    p2 = save("02_broken_path.png", img2)

    # Network of alternatives.
    img3 = base(); d = ImageDraw.Draw(img3)
    nodes = [
        (int(w*0.20), int(h*0.28)), (int(w*0.40), int(h*0.20)), (int(w*0.62), int(h*0.25)),
        (int(w*0.78), int(h*0.43)), (int(w*0.65), int(h*0.66)), (int(w*0.42), int(h*0.72)),
        (int(w*0.20), int(h*0.58)), (int(w*0.50), int(h*0.48)),
    ]
    edges = [(0,1),(1,2),(2,3),(3,4),(4,5),(5,6),(6,0),(0,7),(1,7),(2,7),(3,7),(4,7),(5,7),(6,7)]
    for a,b in edges:
        d.line((nodes[a][0], nodes[a][1], nodes[b][0], nodes[b][1]), fill=dim, width=3)
    for i,(x,y0) in enumerate(nodes):
        r = 34 if i == 7 else 26
        d.ellipse((x-r, y0-r, x+r, y0+r), outline=white, width=3, fill=panel)
    # Reroute highlight.
    for a,b in [(0,7),(7,3),(3,4)]:
        d.line((nodes[a][0], nodes[a][1], nodes[b][0], nodes[b][1]), fill=green, width=max(5, w//360))
    p3 = save("03_network.png", img3)

    # One node fails; network reroutes.
    img4 = img3.copy(); d = ImageDraw.Draw(img4)
    fx, fy = nodes[7]
    d.line((fx-48, fy-48, fx+48, fy+48), fill=red, width=max(7, w//300))
    d.line((fx+48, fy-48, fx-48, fy+48), fill=red, width=max(7, w//300))
    for a,b in [(0,1),(1,2),(2,3),(3,4)]:
        d.line((nodes[a][0], nodes[a][1], nodes[b][0], nodes[b][1]), fill=green, width=max(5, w//360))
    p4 = save("04_reroute.png", img4)

    render_still_schedule([(p1, 3.5), (p2, 3.5), (p3, 4.5), (p4, scene.seconds - 11.5)], out, scene.seconds, fps)
    log(f"Rendered locked alternatives-network clip: {out}")

# -----------------------------------------------------------------------------
# Audio/TTS
# -----------------------------------------------------------------------------

def command_voiceover(args: argparse.Namespace) -> None:
    root = Path(args.root)
    out = root / "audio" / "voiceover.wav"
    ensure_dir(out.parent)
    if args.no_tts:
        log("Skipping TTS generation (--no-tts).")
        return
    generate_tts(out, args.tts_model, args.voice, args.tts_speed)


def generate_tts(out: Path, model: str, voice: str, speed: float) -> None:
    url = f"{API_BASE}/audio/speech"
    payload = {
        "model": model,
        "input": VOICEOVER_SCRIPT,
        "voice": voice,
        "response_format": "wav",
        "speed": speed,
        "instructions": "Urgent but calm documentary narration. Clear, restrained, not ranting. Smart general audience.",
    }
    log(f"Generating voiceover: {out}")
    r = requests.post(url, headers={**openai_headers(), "Content-Type": "application/json"}, json=payload, timeout=300)
    if r.status_code >= 400:
        raise RuntimeError(f"TTS failed: HTTP {r.status_code}\n{r.text}")
    out.write_bytes(r.content)
    log(f"Wrote voiceover: {out} ({ffprobe_duration(out):.1f}s)")


def make_silent_audio(out: Path, seconds: float) -> None:
    check_cmd("ffmpeg")
    ensure_dir(out.parent)
    run_cmd([
        "ffmpeg", "-y",
        "-f", "lavfi",
        "-i", "anullsrc=channel_layout=stereo:sample_rate=48000",
        "-t", f"{seconds:.3f}",
        "-c:a", "pcm_s16le",
        str(out),
    ])

# -----------------------------------------------------------------------------
# Assembly
# -----------------------------------------------------------------------------

def normalize_clip(in_path: Path, out_path: Path, size: str, fps: int, target_seconds: Optional[float] = None) -> None:
    """Normalize a clip and conform it to the scripted scene duration.

    AI context clips are allowed to be shorter than the scripted scene. This
    function loops short clips and trims long clips, strips any generated audio,
    and applies final size/fps.
    """
    check_cmd("ffmpeg")
    w, h = read_size(size)
    ensure_dir(out_path.parent)
    vf = f"scale={w}:{h}:force_original_aspect_ratio=increase,crop={w}:{h},fps={fps},setsar=1,format=yuv420p"
    cmd = ["ffmpeg", "-y"]
    if target_seconds is not None:
        cmd += ["-stream_loop", "-1"]
    cmd += ["-i", str(in_path)]
    if target_seconds is not None:
        cmd += ["-t", f"{target_seconds:.3f}"]
    cmd += [
        "-vf", vf,
        "-an",
        "-c:v", "libx264",
        "-preset", "medium",
        "-crf", "18",
        "-movflags", "+faststart",
        str(out_path),
    ]
    run_cmd(cmd)


def concat_videos(paths: list[Path], out: Path) -> None:
    check_cmd("ffmpeg")
    ensure_dir(out.parent)
    list_file = out.with_suffix(".txt")
    with list_file.open("w", encoding="utf-8") as f:
        for p in paths:
            f.write(f"file '{p.resolve().as_posix()}'\n")
    run_cmd([
        "ffmpeg", "-y",
        "-f", "concat",
        "-safe", "0",
        "-i", str(list_file),
        "-c", "copy",
        str(out),
    ])


def create_end_card(root: Path, size: str, seconds: int, fps: int) -> Path:
    check_cmd("ffmpeg")
    w, h = read_size(size)
    img_path = root / "endcard" / "end_card.png"
    video_path = root / "endcard" / "end_card.mp4"
    ensure_dir(img_path.parent)

    img = Image.new("RGB", (w, h), (4, 4, 7))
    draw = ImageDraw.Draw(img)
    title_font = pil_font(max(42, w // 22), bold=True)
    small_font = pil_font(max(18, w // 70), bold=False)
    tiny_font = pil_font(max(14, w // 90), bold=False)
    white = (245, 245, 245)
    grey = (170, 174, 180)
    red = (205, 60, 60)

    title = "IF THE NEED MATTERS, BUILD ALTERNATIVES"
    bbox = draw.textbbox((0, 0), title, font=title_font)
    draw.text(((w - (bbox[2] - bbox[0])) // 2, int(h * 0.18)), title, font=title_font, fill=white)
    support = "Open source \u2022 Open hardware \u2022 Open standards \u2022 Repair \u2022 Local control"
    bbox2 = draw.textbbox((0, 0), support, font=small_font)
    draw.text(((w - (bbox2[2] - bbox2[0])) // 2, int(h * 0.30)), support, font=small_font, fill=grey)
    draw.line((int(w * 0.22), int(h * 0.38), int(w * 0.78), int(h * 0.38)), fill=red, width=max(3, w // 360))

    logo_files = [
        root / "assets/logos/osi_open_source_initiative.png",
        root / "assets/logos/oshw_logo_filled_color.png",
        root / "assets/logos/beagleboard_official.png",
        # Manual override names:
        root / "assets/logos/osi_uploaded.png",
        root / "assets/logos/oshw_uploaded.png",
        root / "assets/logos/beagleboard_uploaded.png",
    ]
    # Pick at most one for each brand, preferring uploaded override when present.
    brand_paths = [
        root / "assets/logos/osi_uploaded.png" if (root / "assets/logos/osi_uploaded.png").exists() else root / "assets/logos/osi_open_source_initiative.png",
        root / "assets/logos/oshw_uploaded.png" if (root / "assets/logos/oshw_uploaded.png").exists() else root / "assets/logos/oshw_logo_filled_color.png",
        root / "assets/logos/beagleboard_uploaded.png" if (root / "assets/logos/beagleboard_uploaded.png").exists() else root / "assets/logos/beagleboard_official.png",
    ]
    loaded: list[Image.Image] = []
    for p in brand_paths:
        try:
            if p.exists():
                logo = Image.open(p).convert("RGBA")
                loaded.append(logo)
            else:
                loaded.append(None)  # type: ignore[arg-type]
        except Exception as e:
            log(f"WARNING: Could not load logo {p}: {e}")
            loaded.append(None)  # type: ignore[arg-type]

    slot_w = int(w * 0.22)
    slot_h = int(h * 0.13)
    start_x = int(w * 0.14)
    y = int(h * 0.48)
    gap = int(w * 0.06)
    for i, logo in enumerate(loaded):
        x = start_x + i * (slot_w + gap)
        if logo is None:
            # Placeholder if download failed; user should replace with official asset.
            draw.rounded_rectangle((x, y, x + slot_w, y + slot_h), radius=14, outline=(80, 80, 86), width=2)
            placeholder = ["PLACE OFFICIAL", ["OSI", "OSHW", "BEAGLEBOARD"][i], "LOGO HERE"]
            yy = y + 20
            for line in placeholder:
                tb = draw.textbbox((0, 0), line, font=small_font)
                draw.text((x + (slot_w - (tb[2] - tb[0])) // 2, yy), line, font=small_font, fill=grey)
                yy += int(slot_h * 0.28)
            continue
        logo.thumbnail((slot_w, slot_h), Image.LANCZOS)
        lx = x + (slot_w - logo.width) // 2
        ly = y + (slot_h - logo.height) // 2
        # Preserve original logo colors, proportions, shapes, and typography. Composite only.
        img.paste(logo, (lx, ly), logo)

    credit = "Video licensed under Creative Commons [choose license], except third-party trademarks and logos. Logos used for identification; no endorsement implied."
    draw_wrapped(draw, (int(w * 0.08), int(h * 0.78)), credit, tiny_font, grey, int(w * 0.84), line_gap=5)

    img.save(img_path)
    run_cmd([
        "ffmpeg", "-y",
        "-loop", "1",
        "-i", str(img_path),
        "-t", str(seconds),
        "-r", str(fps),
        "-pix_fmt", "yuv420p",
        "-c:v", "libx264",
        "-movflags", "+faststart",
        str(video_path),
    ])
    return video_path


def create_ass_labels(root: Path, out: Path, size: str) -> None:
    w, h = read_size(size)
    font_name = "DejaVu Sans"
    font_size = max(28, int(h * 0.035))
    header = f"""[Script Info]
ScriptType: v4.00+
PlayResX: {w}
PlayResY: {h}
ScaledBorderAndShadow: yes

[V4+ Styles]
Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
Style: Label,{font_name},{font_size},&H00FFFFFF,&H00FFFFFF,&HCC000000,&HAA000000,1,0,0,0,100,100,0,0,1,2,0,7,50,50,50,1
Style: Small,{font_name},{max(20, font_size - 8)},&H00D8D8D8,&H00FFFFFF,&HCC000000,&HAA000000,0,0,0,0,100,100,0,0,1,2,0,7,50,50,50,1

[Events]
Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
"""
    lines = [header]
    t = 0.0
    for scene in SCENES:
        labels = scene.labels or []
        # Scene 1 is rendered locally with the title card already locked in.
        # Do not burn a second WHY DID IT FAIL? subtitle over it.
        if scene.id == 1:
            t += scene.seconds
            continue
        if not labels:
            t += scene.seconds
            continue
        per = max(2.0, scene.seconds / max(1, len(labels)))
        for i, label in enumerate(labels):
            start = t + i * per
            end = min(t + scene.seconds, start + per + 0.25)
            x = int(w * 0.055)
            y = int(h * 0.075) if i % 2 == 0 else int(h * 0.16)
            escaped = ass_escape(label)
            style = "Label" if len(label) < 62 else "Small"
            override = f"{{\\an7\\pos({x},{y})\\bord2\\shad0}}"
            lines.append(f"Dialogue: 0,{seconds_to_ass_time(start)},{seconds_to_ass_time(end)},{style},,0,0,0,,{override}{escaped}\n")
        t += scene.seconds
    # End-card title support is already burned into the end-card image.
    ensure_dir(out.parent)
    out.write_text("".join(lines), encoding="utf-8")
    log(f"Wrote ASS labels: {out}")


def ass_escape(text: str) -> str:
    return text.replace("{", "\\{").replace("}", "\\}").replace("\n", "\\N")


def burn_labels(in_video: Path, ass_file: Path, out_video: Path) -> None:
    check_cmd("ffmpeg")
    ensure_dir(out_video.parent)
    # Escape path for ffmpeg subtitles filter. Forward slashes work on Windows too in many cases.
    ass = ass_file.resolve().as_posix().replace("'", "\\'")
    run_cmd([
        "ffmpeg", "-y",
        "-i", str(in_video),
        "-vf", f"subtitles='{ass}'",
        "-an",
        "-c:v", "libx264",
        "-preset", "medium",
        "-crf", "18",
        "-movflags", "+faststart",
        str(out_video),
    ])


def mux_audio(video: Path, audio: Path, out: Path) -> None:
    check_cmd("ffmpeg")
    ensure_dir(out.parent)
    run_cmd([
        "ffmpeg", "-y",
        "-i", str(video),
        "-i", str(audio),
        "-map", "0:v:0",
        "-map", "1:a:0",
        "-c:v", "copy",
        "-c:a", "aac",
        "-b:a", "192k",
        "-movflags", "+faststart",
        str(out),
    ])


def extend_video_with_endcard(root: Path, base_video: Path, end_card: Path, target_duration: float, size: str, fps: int) -> Path:
    current = ffprobe_duration(base_video)
    if current >= target_duration - 0.25:
        return base_video
    extra = max(1, int(target_duration - current + 1))
    extra_card = create_end_card(root, size, extra, fps)
    out = root / "assembly" / "picture_extended.mp4"
    concat_videos([base_video, extra_card], out)
    return out


def command_assemble(args: argparse.Namespace) -> None:
    root = Path(args.root)
    if getattr(args, "sora", False) and getattr(args, "omni", False):
        raise RuntimeError("Choose only one clip source: --sora or --omni.")
    provider_key = "sora" if getattr(args, "sora", False) else ("gemini_omni" if getattr(args, "omni", False) else None)
    source_dir = context_clips_dir(root, provider_key)
    if not source_dir.exists():
        raise RuntimeError(f"Missing clips directory: {source_dir}. Run generate-clips first.")
    assembly_dir = root / "assembly"
    norm_dir = assembly_dir / "normalized"
    ensure_dir(norm_dir)

    normalized: list[Path] = []
    for scene in SCENES:
        src = source_dir / scene.filename
        if not src.exists():
            raise RuntimeError(f"Missing clip for scene {scene.id}: {src}")
        out = norm_dir / scene.filename
        if not out.exists() or args.force:
            normalize_clip(src, out, args.size, args.fps, target_seconds=scene.seconds)
        normalized.append(out)

    end_card = create_end_card(root, args.size, END_CARD_SECONDS, args.fps)
    picture_no_labels = assembly_dir / "picture_no_labels.mp4"
    concat_videos(normalized + [end_card], picture_no_labels)

    ass_file = assembly_dir / "labels.ass"
    create_ass_labels(root, ass_file, args.size)
    picture_labeled = assembly_dir / "picture_labeled.mp4"
    burn_labels(picture_no_labels, ass_file, picture_labeled)

    audio_path = Path(args.voiceover) if args.voiceover else (root / "audio" / "voiceover.wav")
    if args.tts and not audio_path.exists():
        generate_tts(audio_path, args.tts_model, args.voice, args.tts_speed)
    if not audio_path.exists():
        if args.no_tts:
            duration = ffprobe_duration(picture_labeled)
            audio_path = root / "audio" / "silence.wav"
            make_silent_audio(audio_path, duration)
        else:
            raise RuntimeError(f"Missing voiceover audio: {audio_path}. Use --tts, --no-tts, or --voiceover path.")

    audio_duration = ffprobe_duration(audio_path)
    video_duration = ffprobe_duration(picture_labeled)
    if audio_duration > video_duration + 0.25:
        picture_labeled = extend_video_with_endcard(root, picture_labeled, end_card, audio_duration + 0.5, args.size, args.fps)

    final = root / "exports" / "the_second_question_of_technology.mp4"
    mux_audio(picture_labeled, audio_path, final)
    log(f"Final video: {final}")
    log(f"Duration: video={ffprobe_duration(final):.1f}s")

# -----------------------------------------------------------------------------
# Metadata workbook, CSV, and DDEX draft XML
# -----------------------------------------------------------------------------

def metadata_defaults(root: Path) -> dict[str, str]:
    now = datetime.now(timezone.utc).date().isoformat()
    return {
        "ReleaseTitle": PROJECT_TITLE,
        "VersionTitle": "Original explainer video",
        "DisplayArtist": "BeagleBoard.org Foundation / Open source and open hardware educational project",
        "LabelName": "TBD",
        "ReleaseDate": now,
        "OriginalReleaseDate": now,
        "Genre": "Educational",
        "SubGenre": "Technology / Society",
        "Language": "en",
        "ParentalWarning": "NotExplicit",
        "Duration": "PT3M02S",
        "VideoFile": "exports/the_second_question_of_technology.mp4",
        "Territory": "Worldwide",
        "License": "Creative Commons [choose license], except third-party trademarks and logos",
        "CopyrightText": "(C) [year] [creator]. Some trademarks/logos owned by their respective owners.",
        "Publisher": "TBD",
        "ISRC": "TBD",
        "UPC": "TBD",
        "GRid": "TBD",
        "DPIDSender": "PADPIDA0000000000U",  # placeholder
        "DPIDRecipient": "PADPIDA0000000000R",  # placeholder
        "ProfileWarning": "Draft only. Validate against recipient DDEX ERN 4.3.2 profile before delivery.",
    }


def write_sheet(ws, rows: list[list[Any]], freeze: bool = True) -> None:
    for row in rows:
        ws.append(row)
    if freeze:
        ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="CBD5E1")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(border_style="hair", color="E2E8F0"))
    autofit(ws)


def autofit(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(val) + 2))
        ws.column_dimensions[letter].width = width


def command_metadata(args: argparse.Namespace) -> None:
    root = Path(args.root)
    ensure_dir(root / "metadata")
    md = metadata_defaults(root)
    write_metadata_workbook(root, md)
    write_metadata_csvs(root, md)
    write_ddex_draft_xml(root, md)


def write_metadata_workbook(root: Path, md: dict[str, str]) -> None:
    out = root / "metadata" / "second_question_metadata.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Release_Metadata")
    write_sheet(ws, [
        ["Field", "Value", "DDEX-ish mapping", "Notes"],
        ["ReleaseTitle", md["ReleaseTitle"], "Release/DisplayTitle", "Main release title."],
        ["VersionTitle", md["VersionTitle"], "Release/VersionTitle", "Use if recipient supports version subtitles."],
        ["DisplayArtist", md["DisplayArtist"], "Release/DisplayArtistName", "Replace with final credited entity."],
        ["LabelName", md["LabelName"], "ReleaseLabelReference", "Required by many delivery workflows."],
        ["ReleaseDate", md["ReleaseDate"], "DealList/ReleaseDate or Release/OriginalReleaseDate", "Check recipient profile."],
        ["OriginalReleaseDate", md["OriginalReleaseDate"], "OriginalReleaseDate", ""],
        ["Genre", md["Genre"], "Genre", "DDEX controlled values can vary by recipient."],
        ["SubGenre", md["SubGenre"], "SubGenre", ""],
        ["Language", md["Language"], "LanguageAndScriptCode", "ISO 639 code."],
        ["ParentalWarning", md["ParentalWarning"], "ParentalWarningType", "Likely NotExplicit."],
        ["Territory", md["Territory"], "DealTerms/TerritoryCode", "Worldwide often represented as Worldwide or territory list depending profile."],
        ["License", md["License"], "UsageRights / proprietary notes", "Choose final Creative Commons license."],
        ["CopyrightText", md["CopyrightText"], "CLine/PLine", "Confirm final rights holder and year."],
        ["ProfileWarning", md["ProfileWarning"], "N/A", "This workbook is a crosswalk, not a certified delivery."],
    ])

    ws = wb.create_sheet("Resource_Metadata")
    write_sheet(ws, [
        ["ResourceReference", "ResourceType", "Title", "ISRC/Identifier", "FilePath", "Duration", "Codec", "Language", "Notes"],
        ["A1", "Video", md["ReleaseTitle"], md["ISRC"], md["VideoFile"], md["Duration"], "H.264/AAC MP4", md["Language"], "Final video resource. Replace duration after final export."],
        ["A2", "Image", "End card / cover artwork TBD", "TBD", "assets/cover/cover.jpg", "", "JPEG/PNG", "", "Add separate cover image if required by distributor/DSP."],
    ])

    ws = wb.create_sheet("Contributors")
    write_sheet(ws, [
        ["PartyReference", "Name", "Role", "DDEX role idea", "Identifier", "Notes"],
        ["P1", "TBD", "Creator / Writer", "ComposerLyricist or Author-style proprietary mapping", "TBD", "DDEX roles are music-industry oriented; confirm recipient mapping for educational video."],
        ["P2", "TBD", "Narrator", "Narrator / AssociatedPerformer if supported", "TBD", "If OpenAI TTS is used, disclose according to platform policy."],
        ["P3", "Configured AI context video generators", "Generative video tools", "Proprietary annotation", "", "Sora/Gemini Omni Flash clips are context visuals only; do not credit as human contributors unless your platform requires AI tool disclosure fields."],
    ])

    ws = wb.create_sheet("Rights_And_License")
    write_sheet(ws, [
        ["Item", "Value", "Source/Authority", "Notes"],
        ["Creator-selected video license", "Creative Commons [choose license]", "Creator decision", "Do not finalize until actual license chosen."],
        ["Third-party trademarks/logos", "Excluded from CC license", "Brand guidelines", "Use for identification only; no endorsement implied."],
        ["OSI logo", "Trademarked", "https://opensource.org/about/brand-and-trademark-guidelines", "Use official asset; include attribution/trademark statement as needed."],
        ["Open Source Hardware logo", "CC-SA gear logo; conditions apply", "https://oshwa.org/resources/open-source-hardware-logo/", "Use only as appropriate; avoid certification mark unless project is certified."],
        ["BeagleBoard.org logo", "Trademarked", "https://www.beagleboard.org/brand-use", "Use official asset; do not alter or imply endorsement."],
    ])

    ws = wb.create_sheet("Asset_Manifest")
    rows = [["Key", "Category", "Path", "URL", "Required", "Notes", "SHA256 if downloaded"]]
    for a in DOWNLOAD_ASSETS:
        p = root / a.dest
        rows.append([a.key, a.category, a.dest, a.url, a.required, a.notes, sha256_file(p) if p.exists() else ""])
    write_sheet(ws, rows)

    ws = wb.create_sheet("Sources")
    rows = [["Key", "Title", "URL", "Claim", "Use in video", "License note"]]
    for s in SOURCES:
        rows.append([s.key, s.title, s.url, s.claim, s.use_in_video, s.license_note])
    write_sheet(ws, rows)

    ws = wb.create_sheet("Scenes")
    rows = [["Scene", "Slug", "Seconds", "Voiceover", "Labels", "Source Keys"]]
    for sc in SCENES:
        rows.append([sc.id, sc.slug, sc.seconds, sc.voiceover, "\n".join(sc.labels), ", ".join(sc.source_keys)])
    write_sheet(ws, rows)

    ws = wb.create_sheet("DDEX_Map")
    write_sheet(ws, [
        ["DDEX ERN concept", "Workbook field(s)", "Status", "Notes"],
        ["MessageHeader", "DPIDSender, DPIDRecipient, MessageCreatedDate", "placeholder", "Real DPID values are required for delivery."],
        ["PartyList", "Contributors", "draft", "Parties referenced from resources/releases."],
        ["ResourceList", "Resource_Metadata", "draft", "Video and image resources."],
        ["ReleaseList", "Release_Metadata", "draft", "One release only."],
        ["DealList", "Territory, ReleaseDate", "placeholder", "Recipient-specific commercial terms required."],
        ["TechnicalResourceDetails", "FilePath, Codec, Duration", "draft", "Replace after final mezzanine/export."],
    ])

    ws = wb.create_sheet("Delivery_Checklist")
    write_sheet(ws, [
        ["Check", "Status", "Owner", "Notes"],
        ["Choose final CC license", "Open", "Creator", "Fill end-card credit and workbook field."],
        ["Confirm all logos are official and unmodified", "Open", "Producer", "Downloaded official assets or manually uploaded official files."],
        ["Confirm no certification mark misuse", "Open", "Producer", "Do not use OSHWA certification mark without project UID/certification."],
        ["Final video duration within 2:45-3:20", "Open", "Editor", "Run ffprobe on final export."],
        ["Final subtitles/labels proofread", "Open", "Editor", "Especially Amazon: 'lawsuit alleges'."],
        ["DDEX recipient profile selected", "Open", "Distributor", "Required before real DDEX XML validation."],
        ["ISRC/UPC/GRid assigned as needed", "Open", "Label/Distributor", "Replace TBD values."],
        ["DDEX XML validated", "Open", "Distributor", "Use official/recipient XSD + business rules."],
    ])

    wb.save(out)
    log(f"Wrote metadata workbook: {out}")


def write_metadata_csvs(root: Path, md: dict[str, str]) -> None:
    out_dir = root / "metadata"
    ensure_dir(out_dir)
    release_csv = out_dir / "release_metadata_flat.csv"
    with release_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["field", "value"])
        for k, v in md.items():
            w.writerow([k, v])
    sources_csv = out_dir / "sources_flat.csv"
    with sources_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["key", "title", "url", "claim", "use_in_video", "license_note"])
        w.writeheader()
        for s in SOURCES:
            w.writerow(asdict(s))
    scenes_csv = out_dir / "scenes_flat.csv"
    with scenes_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["id", "slug", "seconds", "voiceover", "labels", "source_keys"])
        w.writeheader()
        for sc in SCENES:
            w.writerow({
                "id": sc.id,
                "slug": sc.slug,
                "seconds": sc.seconds,
                "voiceover": sc.voiceover,
                "labels": " | ".join(sc.labels),
                "source_keys": " | ".join(sc.source_keys),
            })
    log(f"Wrote metadata CSVs in {out_dir}")


def write_ddex_draft_xml(root: Path, md: dict[str, str]) -> None:
    """Write an ERN-style draft XML crosswalk.

    This intentionally uses a simple namespace and conservative placeholders. A real
    DDEX ERN 4.3.2 delivery must use the exact schema, release profile, DPID, identifiers,
    file hashes, deal terms, and recipient business rules.
    """
    out = root / "metadata" / "ern43_draft_not_validated.xml"
    ensure_dir(out.parent)
    ns = "http://ddex.net/xml/ern/43"
    ET.register_namespace("ern", ns)
    root_el = ET.Element(f"{{{ns}}}NewReleaseMessage", attrib={"MessageSchemaVersionId": "ern/43", "LanguageAndScriptCode": md["Language"]})
    header = ET.SubElement(root_el, f"{{{ns}}}MessageHeader")
    ET.SubElement(header, f"{{{ns}}}MessageId").text = f"MSG-{PROJECT_SLUG}-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
    ET.SubElement(header, f"{{{ns}}}MessageCreatedDateTime").text = datetime.now(timezone.utc).isoformat()
    sender = ET.SubElement(header, f"{{{ns}}}MessageSender")
    ET.SubElement(sender, f"{{{ns}}}PartyId").text = md["DPIDSender"]
    ET.SubElement(sender, f"{{{ns}}}PartyName").text = md["LabelName"]
    recipient = ET.SubElement(header, f"{{{ns}}}MessageRecipient")
    ET.SubElement(recipient, f"{{{ns}}}PartyId").text = md["DPIDRecipient"]
    ET.SubElement(recipient, f"{{{ns}}}PartyName").text = "TBD recipient DSP/distributor"

    party_list = ET.SubElement(root_el, f"{{{ns}}}PartyList")
    party = ET.SubElement(party_list, f"{{{ns}}}Party", attrib={"PartyReference": "P1"})
    ET.SubElement(party, f"{{{ns}}}PartyName").text = md["DisplayArtist"]

    resource_list = ET.SubElement(root_el, f"{{{ns}}}ResourceList")
    video = ET.SubElement(resource_list, f"{{{ns}}}Video", attrib={"ResourceReference": "A1"})
    ET.SubElement(video, f"{{{ns}}}ResourceType").text = "Video"
    ET.SubElement(video, f"{{{ns}}}Title").text = md["ReleaseTitle"]
    ET.SubElement(video, f"{{{ns}}}Duration").text = md["Duration"]
    ET.SubElement(video, f"{{{ns}}}ISRC").text = md["ISRC"]
    tech = ET.SubElement(video, f"{{{ns}}}TechnicalDetails")
    ET.SubElement(tech, f"{{{ns}}}FilePath").text = md["VideoFile"]
    ET.SubElement(tech, f"{{{ns}}}CodecType").text = "H.264/AAC MP4"

    image = ET.SubElement(resource_list, f"{{{ns}}}Image", attrib={"ResourceReference": "A2"})
    ET.SubElement(image, f"{{{ns}}}ResourceType").text = "FrontCoverImage"
    ET.SubElement(image, f"{{{ns}}}Title").text = "Cover image TBD"
    ET.SubElement(image, f"{{{ns}}}FilePath").text = "assets/cover/cover.jpg"

    release_list = ET.SubElement(root_el, f"{{{ns}}}ReleaseList")
    release = ET.SubElement(release_list, f"{{{ns}}}Release", attrib={"ReleaseReference": "R1"})
    ET.SubElement(release, f"{{{ns}}}ReleaseType").text = "VideoSingle"
    ET.SubElement(release, f"{{{ns}}}DisplayTitle").text = md["ReleaseTitle"]
    ET.SubElement(release, f"{{{ns}}}DisplayArtistName").text = md["DisplayArtist"]
    ET.SubElement(release, f"{{{ns}}}OriginalReleaseDate").text = md["OriginalReleaseDate"]
    ET.SubElement(release, f"{{{ns}}}ParentalWarningType").text = md["ParentalWarning"]
    ET.SubElement(release, f"{{{ns}}}Genre").text = md["Genre"]
    ET.SubElement(release, f"{{{ns}}}ResourceGroup").text = "A1"

    deal_list = ET.SubElement(root_el, f"{{{ns}}}DealList")
    deal = ET.SubElement(deal_list, f"{{{ns}}}ReleaseDeal", attrib={"ReleaseReference": "R1"})
    terms = ET.SubElement(deal, f"{{{ns}}}DealTerms")
    ET.SubElement(terms, f"{{{ns}}}TerritoryCode").text = md["Territory"]
    ET.SubElement(terms, f"{{{ns}}}CommercialModelType").text = "FreeOfChargeModel"
    ET.SubElement(terms, f"{{{ns}}}UseType").text = "OnDemandStream"
    ET.SubElement(terms, f"{{{ns}}}ValidityPeriodStartDate").text = md["ReleaseDate"]
    ET.SubElement(terms, f"{{{ns}}}LicenseNote").text = md["License"]

    comment = ET.Comment("Draft ERN-style crosswalk only. Validate against actual DDEX ERN 4.3.2 schema, release profile, and recipient business rules before delivery.")
    root_el.insert(0, comment)
    tree = ET.ElementTree(root_el)
    ET.indent(tree, space="  ", level=0)
    tree.write(out, encoding="utf-8", xml_declaration=True)
    log(f"Wrote DDEX draft XML: {out}")

# -----------------------------------------------------------------------------
# All-in-one
# -----------------------------------------------------------------------------

def command_all(args: argparse.Namespace) -> None:
    # Reuse args by building namespace objects for subcommands.
    command_download_assets(args)
    command_metadata(args)
    gen_args = argparse.Namespace(**vars(args))
    gen_args.sora = (not args.mock and DEFAULT_FULL_AI_PROFILE == "sora")
    gen_args.omni = (not args.mock and DEFAULT_FULL_AI_PROFILE == "gemini_omni")
    gen_args.dry_run = False
    command_generate_clips(gen_args)
    asm_args = argparse.Namespace(**vars(args))
    asm_args.sora = (not args.mock and DEFAULT_FULL_AI_PROFILE == "sora")
    asm_args.omni = (not args.mock and DEFAULT_FULL_AI_PROFILE == "gemini_omni")
    asm_args.tts = not args.no_tts
    command_assemble(asm_args)

# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

def add_common(p: argparse.ArgumentParser) -> None:
    p.add_argument("--root", default=str(DEFAULT_ROOT), help="Build directory.")
    p.add_argument("--size", default=DEFAULT_VIDEO_SIZE, help="Video size, e.g. 1920x1080 or 1280x720.")
    p.add_argument("--fps", type=int, default=DEFAULT_FPS)
    p.add_argument("--force", action="store_true", help="Regenerate outputs that already exist.")


def add_context_generation_timing(p: argparse.ArgumentParser) -> None:
    # Model names are intentionally not CLI options. Edit
    # CONTEXT_GENERATION_PROFILES near the top of this file.
    p.add_argument("--poll-seconds", type=int, default=15)
    p.add_argument("--timeout-minutes", type=int, default=90)


def add_tts(p: argparse.ArgumentParser) -> None:
    p.add_argument("--tts", action="store_true", help="Generate/use OpenAI TTS voiceover.")
    p.add_argument("--no-tts", action="store_true", help="Use silence if no voiceover file exists.")
    p.add_argument("--voiceover", default="", help="Path to pre-recorded voiceover WAV/MP3.")
    p.add_argument("--tts-model", default="gpt-4o-mini-tts")
    p.add_argument("--voice", default="alloy")
    p.add_argument("--tts-speed", type=float, default=1.06, help="Voiceover speed; 1.06 helps keep cut under ~3:20.")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build 'The Second Question of Technology' video and metadata.")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("download-assets", help="Download public-domain/source/logo assets and source audit pages.")
    add_common(p)
    p.add_argument("--no-source-pages", action="store_true", help="Skip downloading audit/source pages; still writes source manifest.")
    p.set_defaults(func=command_download_assets)

    p = sub.add_parser("generate-clips", help="Generate AI context clips or local mock clips.")
    add_common(p)
    add_context_generation_timing(p)
    p.add_argument("--sora", action="store_true", help="Use the Sora context profile from CONTEXT_GENERATION_PROFILES.")
    p.add_argument("--omni", action="store_true", help="Use the Gemini Omni Flash context profile from CONTEXT_GENERATION_PROFILES.")
    p.add_argument("--dry-run", action="store_true", help="Write provider request JSONL only; do not call an API.")
    p.add_argument("--only-scene", type=int, default=0, help="Generate only one scene by numeric id; 0 means all scenes.")
    p.set_defaults(func=command_generate_clips)

    p = sub.add_parser("compare-context-generators", help="Generate configured test scenes with both Sora and Gemini Omni Flash.")
    add_common(p)
    add_context_generation_timing(p)
    p.add_argument("--dry-run", action="store_true", help="Write provider request JSONL only; do not call APIs.")
    p.set_defaults(func=command_compare_context_generators)

    p = sub.add_parser("voiceover", help="Generate OpenAI TTS voiceover.")
    add_common(p)
    add_tts(p)
    p.set_defaults(func=command_voiceover)

    p = sub.add_parser("assemble", help="Assemble clips, labels, end card, and audio into final MP4.")
    add_common(p)
    add_tts(p)
    p.add_argument("--sora", action="store_true", help="Use clips from clips/sora instead of clips/mock.")
    p.add_argument("--omni", action="store_true", help="Use clips from clips/omni instead of clips/mock.")
    p.set_defaults(func=command_assemble)

    p = sub.add_parser("metadata", help="Write XLSX, CSV, and ERN-style draft XML metadata.")
    add_common(p)
    p.set_defaults(func=command_metadata)

    p = sub.add_parser("all", help="Run assets, metadata, clips, and assembly.")
    add_common(p)
    add_context_generation_timing(p)
    add_tts(p)
    p.add_argument("--mock", action="store_true", help="Use local mock clips instead of Sora.")
    p.add_argument("--no-source-pages", action="store_true", help="Skip downloading audit/source pages.")
    p.set_defaults(func=command_all)

    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
        return 0
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
